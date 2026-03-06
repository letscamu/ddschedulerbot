"""
DynaBot - Flask Web Application
With authentication and production deployment support
"""

import os
import sys
from datetime import datetime
from functools import wraps

import json

from flask import Flask, render_template, request, jsonify, send_file, redirect, url_for, flash, session
from flask_cors import CORS
from flask_login import LoginManager, UserMixin, login_user, logout_user, login_required, current_user
from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash, check_password_hash
from dotenv import load_dotenv

# Load environment variables from .env file
load_dotenv(os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', '.env'))

# Add backend to path for imports
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from user_store import UserStore, User, VALID_ROLES
from data_loader import DataLoader
from algorithms.des_scheduler import DESScheduler, DayShiftConfig
from exporters.excel_exporter import (
    export_master_schedule,
    export_blast_schedule,
    export_core_schedule,
    export_pending_core_report
)
from exporters.impact_analysis_exporter import generate_impact_analysis
from exporters.resource_utilization_exporter import export_resource_utilization
import gcs_storage


# ============== App Configuration ==============

def create_app():
    """Application factory for Flask app."""
    app = Flask(__name__)

    # Load configuration from environment
    app.secret_key = os.environ.get('SECRET_KEY', 'dev-secret-key-change-in-production')
    app.config['ENV'] = os.environ.get('FLASK_ENV', 'development')
    app.config['DEBUG'] = os.environ.get('FLASK_DEBUG', 'true').lower() == 'true'

    # Session configuration for security
    app.config['SESSION_COOKIE_SECURE'] = os.environ.get('BEHIND_PROXY', 'false').lower() == 'true'
    app.config['SESSION_COOKIE_HTTPONLY'] = True
    app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'

    # File upload configuration
    base_dir = os.path.dirname(os.path.abspath(__file__))
    app.config['UPLOAD_FOLDER'] = os.path.join(base_dir, '..', 'Scheduler Bot Info')
    app.config['OUTPUT_FOLDER'] = os.path.join(base_dir, '..', 'outputs')
    app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 50MB max file size
    app.config['ALLOWED_EXTENSIONS'] = {'xlsx', 'xls'}

    # Ensure directories exist
    os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
    os.makedirs(app.config['OUTPUT_FOLDER'], exist_ok=True)

    # CORS for API access
    CORS(app)

    return app


app = create_app()

# ============== User Management ==============

login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'
login_manager.login_message = 'Please log in to access this page.'
login_manager.login_message_category = 'warning'

# Initialize GCS-backed user store
user_store = UserStore()

# Try loading from GCS first; if no users exist, seed from env vars
if not user_store.load() or len(user_store.list_users()) == 0:
    admin_username = os.environ.get('ADMIN_USERNAME', 'admin')
    admin_password = os.environ.get('ADMIN_PASSWORD', 'admin')
    users_env = os.environ.get('USERS', '')
    user_store.seed_from_env(admin_username, admin_password, users_env)
    print("[Startup] Seeded users from environment variables")


@login_manager.user_loader
def load_user(user_id):
    """Load user by ID for Flask-Login. Only returns active users."""
    return user_store.get_active(user_id)


# ============== Global State ==============

current_schedule = {
    'orders': [],
    'baseline_orders': [],
    'generated_at': None,
    'reports': {},
    'stats': {},
    'is_published': False  # MVP 1.1: Draft until published by Planner/Admin
}

# MVP 1.1: Published schedule (separate from current draft)
published_schedule = {
    'serialized_orders': [],
    'serialized_orders_5day': [],
    'stats': {},
    'stats_5day': {},
    'reports': {},
    'generated_at': None,
    'published_at': None,
    'published_by': None
}

# Planner workflow state — holds draft scenarios and in-progress schedule
planner_state = {
    'scenarios': {},          # Results of 3-scenario simulation
    'base_scenario': None,    # Which scenario was selected as base (e.g., '4day_12h')
    'base_schedule': None,    # The base schedule result (without special requests)
    'final_schedule': None,   # Schedule with approved requests applied
    'simulated_at': None,     # When scenarios were last simulated
    'loader': None,           # DataLoader instance (kept for re-running with requests)
}

# Published schedule — the "official" working schedule visible to all users
published_schedule = {
    'schedule_data': None,    # The published schedule orders/stats
    'published_at': None,
    'published_by': None,
    'mode_label': None,       # e.g., '4 Days x 12 Hours'
}


def load_persisted_schedule():
    """Load schedule state from GCS on startup."""
    global current_schedule, published_schedule
    try:
        state = gcs_storage.load_schedule_state()
        if state:
            current_schedule['generated_at'] = datetime.fromisoformat(state['generated_at']) if state.get('generated_at') else None
            current_schedule['published_by'] = state.get('published_by', '')

            # Load dual-mode data if available, otherwise fall back to single mode
            if state.get('modes'):
                current_schedule['active_mode'] = state.get('active_mode', '4day')
                current_schedule['modes'] = {}
                for mode_key in state['modes']:
                    mode_data = state['modes'][mode_key]
                    current_schedule['modes'][mode_key] = {
                        'serialized_orders': mode_data.get('orders', []),
                        'stats': mode_data.get('stats', {}),
                        'reports': mode_data.get('reports', {})
                    }
                # Backward compat: point to active mode data
                active_mode = state.get('active_mode', '4day')
                active = current_schedule['modes'].get(active_mode, {})
                if not active:
                    # Fallback to first available mode
                    active = next(iter(current_schedule['modes'].values()), {})
                current_schedule['serialized_orders'] = active.get('serialized_orders', [])
                current_schedule['stats'] = active.get('stats', {})
                current_schedule['reports'] = active.get('reports', {})
            else:
                # Legacy single-mode format
                current_schedule['stats'] = state.get('stats', {})
                current_schedule['reports'] = state.get('reports', {})
                current_schedule['serialized_orders'] = state.get('orders', [])

            print(f"[Startup] Loaded persisted schedule with {len(current_schedule.get('serialized_orders', []))} orders")

        # Also load published schedule state
        pub_state = gcs_storage.load_published_schedule()
        if pub_state:
            published_schedule['schedule_data'] = pub_state
            published_schedule['published_at'] = datetime.fromisoformat(pub_state['published_at']) if pub_state.get('published_at') else None
            published_schedule['published_by'] = pub_state.get('published_by')
            published_schedule['mode_label'] = pub_state.get('mode_label')
            print(f"[Startup] Loaded published schedule: {pub_state.get('mode_label')} by {pub_state.get('published_by')}")

    except Exception as e:
        print(f"[Startup] Failed to load persisted schedule: {e}")


# Load persisted schedule on module import
load_persisted_schedule()


# ============== Helper Functions ==============

def allowed_file(filename):
    """Check if file extension is allowed."""
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in app.config['ALLOWED_EXTENSIONS']


def get_uploaded_files():
    """Get list of uploaded files from GCS bucket."""
    try:
        return gcs_storage.get_uploaded_files_info()
    except Exception as e:
        print(f"[WARN] Failed to get files from GCS: {e}")
        # Return empty structure on error
        return {
            'sales_order': None,
            'shop_dispatch': None,
            'hot_list': None,
            'core_mapping': None,
            'process_map': None
        }


def get_available_reports():
    """Get list of generated report files from GCS."""
    try:
        files = gcs_storage.list_files(gcs_storage.OUTPUTS_FOLDER)
    except Exception as e:
        print(f"[WARN] Failed to list reports from GCS: {e}")
        return []

    reports = []
    for file_info in files:
        filename = file_info['name']
        if not filename.endswith('.xlsx') or filename.startswith('~$'):
            continue

        # Determine report type
        report_type = 'Unknown'
        if 'Master_Schedule' in filename:
            report_type = 'Master Schedule'
        elif 'BLAST_Schedule' in filename:
            report_type = 'BLAST Schedule'
        elif 'Core_Oven' in filename or 'Core_Schedule' in filename:
            report_type = 'Core Oven Schedule'
        elif 'Pending_Core' in filename:
            report_type = 'Pending Core Report'
        elif 'Impact_Analysis' in filename:
            report_type = 'Impact Analysis'
        elif 'Resource_Utilization' in filename:
            report_type = 'Resource Utilization'

        reports.append({
            'filename': filename,
            'type': report_type,
            'modified': file_info['modified'],
            'size': file_info['size']
        })

    # Already sorted by modified (newest first) from GCS
    return reports[:50]  # Return most recent 50


# ============== Authentication Routes ==============

@app.route('/login', methods=['GET', 'POST'])
def login():
    """Login page and handler."""
    if current_user.is_authenticated:
        return redirect(url_for('index'))

    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        remember = request.form.get('remember', False)

        user = user_store.get(username)
        if user and user.active and user.check_password(password):
            login_user(user, remember=remember)
            flash(f'Welcome back, {username}!', 'success')

            # Redirect to requested page or dashboard
            next_page = request.args.get('next')
            if next_page and next_page.startswith('/'):
                return redirect(next_page)
            return redirect(url_for('index'))

        flash('Invalid username or password.', 'danger')

    return render_template('login.html')


@app.route('/logout')
@login_required
def logout():
    """Log out the current user."""
    logout_user()
    flash('You have been logged out.', 'info')
    return redirect(url_for('login'))


# ============== Page Routes ==============

@app.route('/')
@login_required
def index():
    """Dashboard home page."""
    files = get_uploaded_files()
    reports = get_available_reports()

    # Load alert summary for dashboard cards
    alerts_data = gcs_storage.load_alerts()
    alert_summary = alerts_data.get('summary', {}) if alerts_data else {}
    alert_generated_at = alerts_data.get('generated_at') if alerts_data else None

    can_generate = current_user.role in ('admin', 'planner')
    return render_template('index.html',
                           files=files,
                           reports=reports[:5],
                           schedule=current_schedule,
                           can_generate=can_generate,
                           alert_summary=alert_summary,
                           alert_generated_at=alert_generated_at)


@app.route('/upload')
@login_required
def upload_page():
    """File upload page."""
    files = get_uploaded_files()
    return render_template('upload.html', files=files)


@app.route('/schedule')
@login_required
def schedule_page():
    """Schedule viewer page."""
    can_generate = current_user.role in ('admin', 'planner')
    return render_template('schedule.html', schedule=current_schedule, can_generate=can_generate)


@app.route('/reports')
@login_required
def reports_page():
    """Reports page."""
    reports = get_available_reports()
    return render_template('reports.html', reports=reports)


@app.route('/simulation')
@login_required
def simulation_page():
    """Visual simulation page."""
    return render_template('simulation.html')


@app.route('/planner')
@login_required
def planner_page():
    """Planner workflow page — scenario comparison, request review, publish."""
    if current_user.role not in ('admin', 'planner'):
        flash('Only Planner and Admin users can access the planner workflow.', 'danger')
        return redirect(url_for('index'))

    # Load pending special requests
    special_requests = gcs_storage.load_special_requests()
    pending_requests = [r for r in special_requests if r.get('status') == 'pending']

    return render_template('planner.html',
                           planner_state=planner_state,
                           published_schedule=published_schedule,
                           pending_request_count=len(pending_requests))


@app.route('/special-requests')
@login_required
def special_requests_page():
    """Dedicated Special Requests page — submit, review, approve/reject."""
    user_can_approve = current_user.role in ('admin', 'planner')
    return render_template('special_requests.html', user_can_approve=user_can_approve)


@app.route('/updates')
@login_required
def update_log_page():
    """Update log and feedback page."""
    return render_template('update_log.html')


@app.route('/mfg-eng-review')
@login_required
def mfg_eng_review_page():
    """Manufacturing Engineering Review page."""
    return render_template('mfg_eng_review.html')


@app.route('/user-management')
@login_required
def user_management_page():
    """User management page — admin only."""
    if current_user.role != 'admin':
        flash('Only administrators can access user management.', 'danger')
        return redirect(url_for('index'))
    return render_template('user_management.html', valid_roles=VALID_ROLES)


@app.route('/core-mapping')
@login_required
def core_mapping_page():
    """Core Mapping read-only view — admin, mfgeng, planner."""
    if current_user.role not in ('admin', 'mfgeng', 'planner'):
        flash('You do not have permission to access Core Mapping.', 'danger')
        return redirect(url_for('index'))
    return render_template('core_mapping.html')


# ============== User Management API ==============


@app.route('/api/users', methods=['GET'])
@login_required
def api_list_users():
    """List all users (admin only)."""
    if current_user.role != 'admin':
        return jsonify({'error': 'Unauthorized'}), 403
    return jsonify({'users': user_store.list_users()})


@app.route('/api/users', methods=['POST'])
@login_required
def api_create_user():
    """Create a new user (admin only)."""
    if current_user.role != 'admin':
        return jsonify({'error': 'Unauthorized'}), 403

    data = request.get_json()
    if not data:
        return jsonify({'error': 'Request body is required.'}), 400

    username = data.get('username', '').strip()
    password = data.get('password', '')
    role = data.get('role', 'guest').strip().lower()

    success, message = user_store.add_user(username, password, role)
    if success:
        return jsonify({'message': message}), 201
    return jsonify({'error': message}), 400


@app.route('/api/users/<username>/role', methods=['PUT'])
@login_required
def api_update_role(username):
    """Update a user's role (admin only)."""
    if current_user.role != 'admin':
        return jsonify({'error': 'Unauthorized'}), 403

    data = request.get_json()
    if not data or 'role' not in data:
        return jsonify({'error': 'Role is required.'}), 400

    success, message = user_store.update_role(username, data['role'].strip().lower())
    if success:
        return jsonify({'message': message})
    return jsonify({'error': message}), 400


@app.route('/api/users/<username>/reset-password', methods=['PUT'])
@login_required
def api_reset_password(username):
    """Admin password reset (admin only)."""
    if current_user.role != 'admin':
        return jsonify({'error': 'Unauthorized'}), 403

    data = request.get_json()
    if not data or 'password' not in data:
        return jsonify({'error': 'New password is required.'}), 400

    success, message = user_store.reset_password(username, data['password'])
    if success:
        return jsonify({'message': message})
    return jsonify({'error': message}), 400


@app.route('/api/users/<username>/disable', methods=['PUT'])
@login_required
def api_disable_user(username):
    """Disable a user account (admin only)."""
    if current_user.role != 'admin':
        return jsonify({'error': 'Unauthorized'}), 403

    if username == current_user.username:
        return jsonify({'error': 'You cannot disable your own account.'}), 400

    success, message = user_store.disable_user(username)
    if success:
        return jsonify({'message': message})
    return jsonify({'error': message}), 400


@app.route('/api/users/<username>/enable', methods=['PUT'])
@login_required
def api_enable_user(username):
    """Re-enable a disabled user account (admin only)."""
    if current_user.role != 'admin':
        return jsonify({'error': 'Unauthorized'}), 403

    success, message = user_store.enable_user(username)
    if success:
        return jsonify({'message': message})
    return jsonify({'error': message}), 400


@app.route('/api/users/me/password', methods=['PUT'])
@login_required
def api_change_own_password():
    """Self-service password change (requires current password)."""
    data = request.get_json()
    if not data:
        return jsonify({'error': 'Request body is required.'}), 400

    current_password = data.get('current_password', '')
    new_password = data.get('new_password', '')

    if not current_password or not new_password:
        return jsonify({'error': 'Current and new passwords are required.'}), 400

    success, message = user_store.change_password(
        current_user.username, current_password, new_password
    )
    if success:
        return jsonify({'message': message})
    return jsonify({'error': message}), 400


# ============== Core Mapping API ==============


@app.route('/api/core-mapping')
@login_required
def api_core_mapping():
    """Get core mapping data for the read-only view."""
    if current_user.role not in ('admin', 'mfgeng', 'planner'):
        return jsonify({'error': 'Unauthorized'}), 403

    try:
        # Download files and load data
        import tempfile
        temp_dir = tempfile.mkdtemp(prefix='estradabot_core_')
        local_paths = gcs_storage.download_files_for_processing(temp_dir)

        loader = DataLoader(local_paths)

        # Build mapping records
        mapping_records = []
        for part_number, data in loader.core_mapping.items():
            mapping_records.append({
                'part_number': part_number,
                'description': data.get('description', ''),
                'core_number': data.get('core_number'),
                'rubber_type': data.get('rubber_type', ''),
                'injection_time': data.get('injection_time'),
                'cure_time': data.get('cure_time'),
                'quench_time': data.get('quench_time'),
                'stator_od': data.get('stator_od'),
                'lobe_config': data.get('lobe_config'),
                'stage_count': data.get('stage_count'),
                'fit': data.get('fit', ''),
            })

        # Build inventory records
        inventory_records = []
        for core_number, cores in loader.core_inventory.items():
            for core in cores:
                inventory_records.append({
                    'core_number': core_number,
                    'suffix': core.get('suffix', ''),
                    'core_pn': core.get('core_pn', ''),
                    'model': core.get('model', ''),
                    'tooling_pn': core.get('tooling_pn', ''),
                    'state': core.get('state', 'available'),
                })

        # Detect mismatches: parts in orders referencing cores not in inventory
        import pandas as pd
        mismatches = []
        inventory_core_set = set(loader.core_inventory.keys())

        # Parts with no core mapping
        unmapped = loader.validation_results.get('unmapped_parts', [])
        for part in unmapped:
            mismatches.append({
                'part_number': part,
                'reason': 'Part in orders but not in core mapping',
            })

        # Parts mapped to cores not in inventory
        for part_number, data in loader.core_mapping.items():
            core_num = data.get('core_number')
            if core_num and not pd.isna(core_num):
                try:
                    core_int = int(float(core_num))
                except (ValueError, TypeError):
                    core_int = core_num
                if core_int not in inventory_core_set:
                    mismatches.append({
                        'part_number': part_number,
                        'core_number': core_num,
                        'reason': 'Core in mapping but not in inventory',
                    })

        # Clean up temp dir
        import shutil
        shutil.rmtree(temp_dir, ignore_errors=True)

        return jsonify({
            'mapping': mapping_records,
            'inventory': inventory_records,
            'mismatches': mismatches,
            'summary': {
                'total_parts_mapped': len(mapping_records),
                'total_core_types': len(loader.core_inventory),
                'total_physical_cores': sum(len(c) for c in loader.core_inventory.values()),
                'mismatch_count': len(mismatches),
            }
        })

    except Exception as e:
        print(f"[Core Mapping API] Error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'Failed to load core mapping data: {str(e)}'}), 500


# ============== Schedule Reorder API ==============


@app.route('/api/schedule/reorder', methods=['POST'])
@login_required
def api_reorder_schedule():
    """Save a manual reorder of the schedule. Admin/planner only.
    Accepts {mode: '4day'|'5day', sequence: ['WO-001', 'WO-002', ...]}
    Recalculates BLAST sequence numbers based on new order."""
    if current_user.role not in ('admin', 'planner'):
        return jsonify({'error': 'Only Planner and Admin users can reorder.'}), 403

    data = request.get_json()
    if not data or 'sequence' not in data:
        return jsonify({'error': 'Sequence array is required.'}), 400

    mode = data.get('mode', '4day')
    sequence = data.get('sequence', [])

    if not sequence:
        return jsonify({'error': 'Sequence cannot be empty.'}), 400

    # Save reorder state to GCS
    reorder_state = {
        'mode': mode,
        'sequence': sequence,
        'created_by': current_user.username,
        'created_at': datetime.now().isoformat(),
    }
    gcs_storage.save_reorder_state(reorder_state)

    return jsonify({
        'message': f'Schedule reordered ({len(sequence)} orders).',
        'sequence': sequence,
    })


@app.route('/api/schedule/reorder', methods=['DELETE'])
@login_required
def api_clear_reorder():
    """Clear custom ordering, reverting to scheduler output. Admin/planner only."""
    if current_user.role not in ('admin', 'planner'):
        return jsonify({'error': 'Unauthorized'}), 403

    gcs_storage.clear_reorder_state()
    return jsonify({'message': 'Custom ordering cleared.'})


@app.route('/api/schedule/reorder/status')
@login_required
def api_reorder_status():
    """Check if custom ordering exists for current schedule."""
    state = gcs_storage.load_reorder_state()
    if state:
        return jsonify({
            'has_reorder': True,
            'mode': state.get('mode'),
            'created_by': state.get('created_by'),
            'created_at': state.get('created_at'),
            'order_count': len(state.get('sequence', [])),
        })
    return jsonify({'has_reorder': False})


# ============== Reconciliation Helpers ==============


def _reconcile_special_requests(filename: str, file_type: str) -> int:
    """
    After a file upload, check for unmatched (Mode B) special requests
    whose WO numbers now appear in the uploaded data.
    Detects data mismatches and flags them for planner review.
    Returns the number of newly matched requests.
    """
    try:
        all_requests = gcs_storage.load_special_requests()
        unmatched = [r for r in all_requests if r.get('matched') is False
                     and r.get('status') == 'pending']
        if not unmatched:
            return 0

        # Extract WO numbers from the uploaded file
        uploaded_orders = {}  # wo_number -> order dict
        import tempfile
        temp_dir = tempfile.mkdtemp(prefix='estradabot_reconcile_')
        local_paths = gcs_storage.download_files_for_processing(temp_dir)

        # Parse orders from all available data
        try:
            loader = DataLoader(local_paths)
            if loader.orders:
                for o in loader.orders:
                    wo = o.get('wo_number')
                    if wo:
                        uploaded_orders[wo] = o
        except Exception as e:
            print(f"[Reconcile] Could not parse uploaded data: {e}")
            return 0

        # Match unmatched requests and check for data mismatches
        matched_count = 0
        for req in all_requests:
            if req.get('matched') is False and req.get('status') == 'pending':
                wo = req['wo_number']
                if wo in uploaded_orders:
                    order_data = uploaded_orders[wo]
                    req['matched'] = True
                    req['matched_at'] = datetime.now().isoformat()
                    matched_count += 1

                    # Check for data mismatches between request and actual order
                    mismatches = []
                    if req.get('part_number') and order_data.get('part_number'):
                        if req['part_number'] != order_data['part_number']:
                            mismatches.append({
                                'field': 'part_number',
                                'expected': req['part_number'],
                                'actual': order_data['part_number']
                            })
                    if req.get('customer') and order_data.get('customer'):
                        if req['customer'].lower() != str(order_data['customer']).lower():
                            mismatches.append({
                                'field': 'customer',
                                'expected': req['customer'],
                                'actual': order_data.get('customer', '')
                            })

                    if mismatches:
                        req['data_mismatches'] = mismatches
                        req['needs_review'] = True
                        print(f"[Reconcile] WARNING: Data mismatch for {req['id']} (WO {wo}): {mismatches}")
                    else:
                        req['data_mismatches'] = []
                        req['needs_review'] = False

                    # Store matched order data for reference
                    req['matched_order_data'] = {
                        'part_number': order_data.get('part_number', ''),
                        'customer': str(order_data.get('customer', '')),
                        'description': order_data.get('description', ''),
                    }

                    print(f"[Reconcile] Matched special request {req['id']} to WO {wo}")

        if matched_count > 0:
            gcs_storage.save_special_requests(all_requests)
            print(f"[Reconcile] {matched_count} request(s) matched from upload of {filename}")

            # Create notification about matched requests
            mismatch_count = sum(1 for r in all_requests
                                if r.get('matched_at') and r.get('needs_review'))
            msg = f'{matched_count} Mode B request(s) matched to uploaded data'
            if mismatch_count:
                msg += f' ({mismatch_count} with data mismatches — review needed)'
            create_notification(
                'warning' if mismatch_count else 'info',
                msg,
                target_roles=['admin', 'planner'],
                related_entity={'type': 'reconciliation', 'value': filename}
            )

        # Clean up temp dir
        import shutil
        shutil.rmtree(temp_dir, ignore_errors=True)

        return matched_count
    except Exception as e:
        print(f"[Reconcile] Error during reconciliation: {e}")
        return 0


# ============== API Routes ==============

@app.route('/api/upload', methods=['POST'])
@login_required
def upload_file():
    """Handle file upload to GCS."""
    if 'file' not in request.files:
        return jsonify({'error': 'No file provided'}), 400

    file = request.files['file']
    file_type = request.form.get('type', 'unknown')

    if file.filename == '':
        return jsonify({'error': 'No file selected'}), 400

    if not allowed_file(file.filename):
        return jsonify({'error': 'Invalid file type. Only .xlsx and .xls files allowed.'}), 400

    # Upload to GCS
    filename = secure_filename(file.filename)
    try:
        # Scrub sensitive columns from sales order files before uploading
        if file_type == 'sales_order' or 'open sales order' in filename.lower().replace('_', ' ') or filename.lower().startswith('oso'):
            import tempfile
            import openpyxl

            # Save to temp file for processing
            fd, temp_path = tempfile.mkstemp(suffix='.xlsx')
            os.close(fd)
            file.save(temp_path)

            # Open and scrub sensitive columns
            wb = openpyxl.load_workbook(temp_path)
            scrubbed_columns = []
            sensitive_headers = ['unit price', 'net price', 'customer address', 'address']

            # Drop all sheets except RawData (SAP exports include many extra tabs)
            target_sheet = 'RawData'
            if target_sheet not in wb.sheetnames:
                wb.close()
                os.unlink(temp_path)
                return jsonify({'error': f'Invalid file: expected a "RawData" sheet but found: {", ".join(wb.sheetnames)}'}), 400
            sheets_to_remove = [s for s in wb.sheetnames if s != target_sheet]
            for sheet_name in sheets_to_remove:
                del wb[sheet_name]
            if sheets_to_remove:
                print(f"[Scrub] Removed {len(sheets_to_remove)} extra sheet(s) from {filename}: {sheets_to_remove}")

            for ws in wb.worksheets:
                cols_to_delete = []
                for col_idx in range(1, ws.max_column + 1):
                    header = ws.cell(row=1, column=col_idx).value
                    if header and str(header).strip().lower() in sensitive_headers:
                        cols_to_delete.append(col_idx)
                        scrubbed_columns.append(str(header).strip())

                # Delete columns in reverse order to preserve indices
                for col_idx in sorted(cols_to_delete, reverse=True):
                    ws.delete_cols(col_idx)

            wb.save(temp_path)
            wb.close()

            if scrubbed_columns:
                print(f"[Scrub] Removed sensitive columns from {filename}: {scrubbed_columns}")

            # Upload scrubbed file to GCS
            gcs_storage.upload_file(temp_path, filename)
            os.unlink(temp_path)
        else:
            gcs_storage.upload_file_object(file, filename)

        # Reconcile unmatched special requests (Mode B placeholders)
        matched_count = _reconcile_special_requests(filename, file_type)

        flash_msg = f'File "{filename}" uploaded successfully!'
        if matched_count > 0:
            flash_msg += f' {matched_count} pending special request(s) matched to new data.'
        flash(flash_msg, 'success')

        return jsonify({
            'success': True,
            'filename': filename,
            'type': file_type,
            'matched_requests': matched_count,
        })
    except Exception as e:
        print(f"[ERROR] Failed to upload to GCS: {e}")
        return jsonify({'error': f'Failed to upload file: {str(e)}'}), 500


def _run_schedule_mode(loader, working_days, mode_label, temp_dir, timestamp,
                       shift_hours=12, skip_hot_list=False, day_configs=None):
    """
    Run the scheduler for a given working_days and shift_hours configuration.
    Returns dict with orders, baseline_orders, reports, stats, serialized_orders.

    Args:
        loader: DataLoader with loaded input data
        working_days: List of weekday ints (e.g., [0,1,2,3] for Mon-Thu)
        mode_label: Label for report filenames (e.g., '4Day_10h')
        temp_dir: Temp directory for report files
        timestamp: Timestamp string for filenames
        shift_hours: 10 or 12 hour shifts. Defaults to 12.
        skip_hot_list: If True, only generate baseline schedule (no hot list processing)
        day_configs: Optional per-day DayShiftConfig dict for advanced mode.
    """
    from datetime import timedelta
    AT_RISK_BUFFER_DAYS = 2

    # Create scheduler
    scheduler = DESScheduler(
        orders=loader.orders,
        core_mapping=loader.core_mapping,
        core_inventory=loader.core_inventory,
        working_days=working_days,
        shift_hours=shift_hours,
        day_configs=day_configs,
        wip_orders=loader.wip_in_process_orders
    )

    # Run baseline schedule (without hot list)
    baseline_orders = scheduler.schedule_orders()
    active_scheduler = scheduler

    # If hot list exists and not skipping, run with hot list
    scheduled_orders = baseline_orders
    if not skip_hot_list and loader.hot_list_entries:
        scheduler_with_hot = DESScheduler(
            orders=loader.orders,
            core_mapping=loader.core_mapping,
            core_inventory=loader.core_inventory,
            working_days=working_days,
            shift_hours=shift_hours,
            day_configs=day_configs,
            wip_orders=loader.wip_in_process_orders
        )
        scheduled_orders = scheduler_with_hot.schedule_orders(
            hot_list_entries=loader.hot_list_entries
        )
        active_scheduler = scheduler_with_hot

    # Export reports
    reports = {}

    master_filename = f'Master_Schedule_{mode_label}_{timestamp}.xlsx'
    master_path = os.path.join(temp_dir, master_filename)
    export_master_schedule(scheduled_orders, master_path)
    gcs_storage.upload_file(master_path, master_filename, gcs_storage.OUTPUTS_FOLDER)
    reports['master'] = master_filename

    blast_filename = f'BLAST_Schedule_{mode_label}_{timestamp}.xlsx'
    blast_path = os.path.join(temp_dir, blast_filename)
    on_blaster = [o for o in loader.wip_in_process_orders if o.get('current_operation') == 1300]
    export_blast_schedule(scheduled_orders, blast_path, currently_blasting=on_blaster)
    gcs_storage.upload_file(blast_path, blast_filename, gcs_storage.OUTPUTS_FOLDER)
    reports['blast'] = blast_filename

    core_filename = f'Core_Oven_Schedule_{mode_label}_{timestamp}.xlsx'
    core_path = os.path.join(temp_dir, core_filename)
    export_core_schedule(scheduled_orders, core_path)
    gcs_storage.upload_file(core_path, core_filename, gcs_storage.OUTPUTS_FOLDER)
    reports['core'] = core_filename

    pending_filename = f'Pending_Core_{mode_label}_{timestamp}.xlsx'
    pending_path = os.path.join(temp_dir, pending_filename)
    pending_orders = getattr(active_scheduler, 'pending_core_orders', [])
    export_pending_core_report(pending_orders, pending_path)
    gcs_storage.upload_file(pending_path, pending_filename, gcs_storage.OUTPUTS_FOLDER)
    reports['pending'] = pending_filename

    utilization_filename = f'Resource_Utilization_{mode_label}_{timestamp}.xlsx'
    utilization_path = os.path.join(temp_dir, utilization_filename)
    export_resource_utilization(scheduled_orders, utilization_path)
    gcs_storage.upload_file(utilization_path, utilization_filename, gcs_storage.OUTPUTS_FOLDER)
    reports['utilization'] = utilization_filename

    # Impact analysis if hot list was used
    if loader.hot_list_entries:
        hot_list_core_shortages = getattr(active_scheduler, 'hot_list_core_shortages', [])
        impact_path = generate_impact_analysis(
            scheduled_orders,
            baseline_orders,
            loader.hot_list_entries,
            hot_list_core_shortages,
            temp_dir
        )
        impact_filename = os.path.basename(impact_path)
        gcs_storage.upload_file(impact_path, impact_filename, gcs_storage.OUTPUTS_FOLDER)
        reports['impact'] = impact_filename

    # Calculate stats
    on_time_count = 0
    late_count = 0
    at_risk_count = 0

    for o in scheduled_orders:
        if not getattr(o, 'on_time', True):
            late_count += 1
        else:
            deadline = getattr(o, 'basic_finish_date', None) or getattr(o, 'promise_date', None)
            completion = getattr(o, 'completion_date', None)
            if deadline and completion:
                days_buffer = (deadline - completion).days
                if days_buffer <= AT_RISK_BUFFER_DAYS:
                    at_risk_count += 1
                else:
                    on_time_count += 1
            else:
                on_time_count += 1

    turnaround_times = [o.turnaround_days for o in scheduled_orders if o.turnaround_days]
    avg_turnaround = sum(turnaround_times) / len(turnaround_times) if turnaround_times else 0

    # Serialize orders
    serialized_orders = []
    for order in scheduled_orders:
        deadline = order.basic_finish_date or order.promise_date
        if not order.on_time:
            status = 'Late'
        elif deadline and order.completion_date:
            days_buffer = (deadline - order.completion_date).days
            status = 'At Risk' if days_buffer <= AT_RISK_BUFFER_DAYS else 'On Time'
        else:
            status = 'On Time'

        serialized_orders.append({
            'wo_number': order.wo_number or '',
            'serial_number': order.serial_number or '',
            'part_number': order.part_number or '',
            'description': order.description or '',
            'customer': order.customer or '',
            'assigned_core': order.assigned_core or '',
            'rubber_type': order.rubber_type or '',
            'priority': order.priority,
            'blast_date': order.blast_date.isoformat() if order.blast_date else None,
            'completion_date': order.completion_date.isoformat() if order.completion_date else None,
            'promise_date': order.promise_date.isoformat() if order.promise_date else None,
            'basic_finish_date': order.basic_finish_date.isoformat() if order.basic_finish_date else None,
            'turnaround_days': order.turnaround_days,
            'on_time': order.on_time,
            'on_time_status': status,
            'is_reline': order.is_reline,
            'special_instructions': order.special_instructions or '',
            'supermarket_location': order.supermarket_location or ''
        })

    stats = {
        'total_orders': len(scheduled_orders),
        'on_time': on_time_count,
        'late': late_count,
        'at_risk': at_risk_count,
        'avg_turnaround': round(avg_turnaround, 1),
        'hot_list_count': len(loader.hot_list_entries) if loader.hot_list_entries else 0
    }

    return {
        'orders': scheduled_orders,
        'baseline_orders': baseline_orders,
        'reports': reports,
        'stats': stats,
        'serialized_orders': serialized_orders
    }


@app.route('/api/generate', methods=['POST'])
@login_required
def generate_schedule():
    """Generate schedule from uploaded files in GCS. Only admin/planner roles.
    Runs both 4-day and 5-day schedules."""
    global current_schedule

    # Role check: only admin and planner can generate schedules
    if current_user.role not in ('admin', 'planner'):
        return jsonify({'error': 'Only Planner and Admin users can generate schedules.'}), 403

    # Check for existing custom reorder and warn
    req_data = request.get_json() or {}
    reorder_state = gcs_storage.load_reorder_state()
    if reorder_state and not req_data.get('confirm_discard_reorder'):
        return jsonify({
            'warning': 'custom_order_exists',
            'message': f'Custom ordering exists (set by {reorder_state.get("created_by", "unknown")}). '
                       f'Generating a new schedule will discard this custom order.',
            'reorder_by': reorder_state.get('created_by'),
            'reorder_at': reorder_state.get('created_at'),
        }), 409

    # Clear reorder state since we're regenerating
    if reorder_state:
        gcs_storage.clear_reorder_state()

    try:
        # Download files from GCS to local temp directory
        import tempfile
        temp_dir = tempfile.mkdtemp(prefix='estradabot_')
        print(f"[Generate] Downloading files from GCS to {temp_dir}")

        local_paths = gcs_storage.download_files_for_processing(temp_dir)
        print(f"[Generate] Downloaded files: {local_paths}")

        # Load data from temp directory
        loader = DataLoader(data_dir=temp_dir)
        loader.load_all()

        if not loader.orders:
            return jsonify({'error': 'No orders loaded. Please upload a Sales Order file.'}), 400

        # Exclude orders on hold
        order_holds = gcs_storage.load_order_holds()
        if order_holds:
            before_count = len(loader.orders)
            loader.orders = [o for o in loader.orders if o.get('wo_number') not in order_holds]
            held_count = before_count - len(loader.orders)
            if held_count > 0:
                print(f"[Generate] Excluded {held_count} orders on hold")

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')

        # Run 4-day 12h schedule (Mon-Thu, default)
        print("[Generate] Running 4-day 12h schedule (Mon-Thu)...")
        result_4day = _run_schedule_mode(loader, [0, 1, 2, 3], '4Day', temp_dir, timestamp, shift_hours=12)

        # Run 5-day 12h schedule (Mon-Fri)
        print("[Generate] Running 5-day 12h schedule (Mon-Fri)...")
        result_5day = _run_schedule_mode(loader, [0, 1, 2, 3, 4], '5Day', temp_dir, timestamp, shift_hours=12)

        # Clean up temp directory
        import shutil
        shutil.rmtree(temp_dir, ignore_errors=True)
        print(f"[Generate] Cleaned up temp directory {temp_dir}")

        generated_at = datetime.now()

        # Update global state with both modes
        current_schedule = {
            'generated_at': generated_at,
            'published_by': current_user.username,
            'active_mode': '4day',
            'modes': {
                '4day': {
                    'orders': result_4day['orders'],
                    'baseline_orders': result_4day['baseline_orders'],
                    'reports': result_4day['reports'],
                    'stats': result_4day['stats'],
                    'serialized_orders': result_4day['serialized_orders']
                },
                '5day': {
                    'orders': result_5day['orders'],
                    'baseline_orders': result_5day['baseline_orders'],
                    'reports': result_5day['reports'],
                    'stats': result_5day['stats'],
                    'serialized_orders': result_5day['serialized_orders']
                }
            },
            # Backward compatibility: point to active mode
            'orders': result_4day['orders'],
            'baseline_orders': result_4day['baseline_orders'],
            'reports': result_4day['reports'],
            'stats': result_4day['stats'],
            'serialized_orders': result_4day['serialized_orders']
        }

        # Persist to GCS
        gcs_storage.save_schedule_state({
            'generated_at': generated_at.isoformat(),
            'published_by': current_user.username,
            'active_mode': '4day',
            'modes': {
                '4day': {
                    'orders': result_4day['serialized_orders'],
                    'stats': result_4day['stats'],
                    'reports': result_4day['reports']
                },
                '5day': {
                    'orders': result_5day['serialized_orders'],
                    'stats': result_5day['stats'],
                    'reports': result_5day['reports']
                }
            },
            # Backward compat
            'orders': result_4day['serialized_orders'],
            'stats': result_4day['stats'],
            'reports': result_4day['reports']
        })

        total_4 = result_4day['stats']['total_orders']
        total_5 = result_5day['stats']['total_orders']
        flash(f'Schedule generated successfully! 4-day: {total_4} orders, 5-day: {total_5} orders.', 'success')
        return jsonify({
            'success': True,
            'stats': result_4day['stats'],
            'stats_5day': result_5day['stats'],
            'reports': result_4day['reports']
        })

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


def _serialize_orders_from_objects(orders, stats_ref):
    """Serialize in-memory ScheduledOrder objects to dicts for the API."""
    AT_RISK_BUFFER_DAYS = 2
    orders_data = []
    on_time_count = 0
    late_count = 0
    at_risk_count = 0

    for order in orders:
        if not order.on_time:
            status = 'Late'
            late_count += 1
        else:
            deadline = order.basic_finish_date or order.promise_date
            if deadline and order.completion_date:
                days_buffer = (deadline - order.completion_date).days
                if days_buffer <= AT_RISK_BUFFER_DAYS:
                    status = 'At Risk'
                    at_risk_count += 1
                else:
                    status = 'On Time'
                    on_time_count += 1
            else:
                status = 'On Time'
                on_time_count += 1

        orders_data.append({
            'wo_number': order.wo_number or '',
            'serial_number': order.serial_number or '',
            'part_number': order.part_number or '',
            'description': order.description or '',
            'customer': order.customer or '',
            'core': order.assigned_core or '',
            'rubber_type': order.rubber_type or '',
            'priority': order.priority,
            'blast_date': order.blast_date.isoformat() if order.blast_date else '',
            'completion_date': order.completion_date.isoformat() if order.completion_date else '',
            'promise_date': order.promise_date.isoformat() if order.promise_date else '',
            'turnaround_days': order.turnaround_days or '',
            'on_time_status': status,
            'is_rework': order.is_reline,
            'special_instructions': order.special_instructions or '',
            'supermarket_location': order.supermarket_location or ''
        })

    turnaround_times = [o.turnaround_days for o in orders if o.turnaround_days]
    avg_turnaround = sum(turnaround_times) / len(turnaround_times) if turnaround_times else 0

    fresh_stats = {
        'total_orders': len(orders_data),
        'on_time': on_time_count,
        'late': late_count,
        'at_risk': at_risk_count,
        'avg_turnaround': round(avg_turnaround, 1),
        'hot_list_count': stats_ref.get('hot_list_count', 0) if stats_ref else 0
    }

    return orders_data, fresh_stats


def _serialize_orders_from_dicts(serialized_orders):
    """Map serialized order dicts to the API response format."""
    orders_data = []
    for order in serialized_orders:
        orders_data.append({
            'wo_number': order.get('wo_number', ''),
            'serial_number': order.get('serial_number', ''),
            'part_number': order.get('part_number', ''),
            'description': order.get('description', ''),
            'customer': order.get('customer', ''),
            'core': order.get('assigned_core', ''),
            'rubber_type': order.get('rubber_type', ''),
            'priority': order.get('priority', ''),
            'blast_date': order.get('blast_date', ''),
            'completion_date': order.get('completion_date', ''),
            'promise_date': order.get('promise_date', ''),
            'turnaround_days': order.get('turnaround_days', ''),
            'on_time_status': order.get('on_time_status', 'On Time'),
            'is_rework': order.get('is_reline', False),
            'special_instructions': order.get('special_instructions', ''),
            'supermarket_location': order.get('supermarket_location', '')
        })
    return orders_data


def _apply_reorder(orders_data, mode):
    """Apply custom reorder sequence to orders list if one exists for this mode."""
    reorder_state = gcs_storage.load_reorder_state()
    if not reorder_state or reorder_state.get('mode') != mode:
        return orders_data, False

    sequence = reorder_state.get('sequence', [])
    if not sequence:
        return orders_data, False

    # Build index: WO# -> order dict
    order_map = {o['wo_number']: o for o in orders_data}

    reordered = []
    for wo in sequence:
        if wo in order_map:
            reordered.append(order_map.pop(wo))

    # Append any orders not in the saved sequence (new orders added after reorder)
    for o in orders_data:
        if o['wo_number'] in order_map:
            reordered.append(o)

    return reordered, True


@app.route('/api/schedule')
@login_required
def get_schedule():
    """Get current schedule data as JSON. Accepts ?mode= query parameter."""
    has_modes = current_schedule.get('modes')
    active_mode = current_schedule.get('active_mode', '4day')
    mode = request.args.get('mode', active_mode)

    # Resolve the data source for the requested mode
    mode_data = None
    if has_modes:
        mode_data = current_schedule['modes'].get(mode)
        # Fallback: if requested mode not found, use active or first available
        if not mode_data:
            mode_data = current_schedule['modes'].get(active_mode)
            if not mode_data and current_schedule['modes']:
                mode = next(iter(current_schedule['modes']))
                mode_data = current_schedule['modes'][mode]

    orders_data = None
    stats = {}
    resp_has_modes = False

    # Case 1: We have in-memory ScheduledOrder objects (just generated)
    if mode_data and mode_data.get('orders'):
        orders_data, stats = _serialize_orders_from_objects(
            mode_data['orders'], mode_data.get('stats', {})
        )
        resp_has_modes = True

    # Case 2: We have serialized orders from mode data (loaded from GCS)
    elif mode_data and mode_data.get('serialized_orders'):
        orders_data = _serialize_orders_from_dicts(mode_data['serialized_orders'])
        stats = mode_data.get('stats', {})
        resp_has_modes = True

    # Case 3: Legacy single-mode data (in-memory objects)
    elif current_schedule.get('orders'):
        orders_data, stats = _serialize_orders_from_objects(
            current_schedule['orders'], current_schedule.get('stats', {})
        )
        mode = '4day'

    # Case 4: Legacy single-mode data (serialized dicts from GCS)
    elif current_schedule.get('serialized_orders'):
        orders_data = _serialize_orders_from_dicts(current_schedule['serialized_orders'])
        stats = current_schedule.get('stats', {})
        mode = '4day'

    if not orders_data:
        return jsonify({'orders': [], 'stats': {}, 'mode': '4day', 'has_modes': False, 'has_reorder': False})

    # Apply custom reorder if it exists for this mode
    orders_data, has_reorder = _apply_reorder(orders_data, mode)

    return jsonify({
        'orders': orders_data,
        'stats': stats,
        'mode': mode,
        'has_modes': resp_has_modes or has_modes is not None,
        'has_reorder': has_reorder,
        'generated_at': current_schedule['generated_at'].isoformat() if current_schedule.get('generated_at') else None,
        'published_by': current_schedule.get('published_by', '')
    })


@app.route('/api/download/<filename>')
@login_required
def download_report(filename):
    """Download a report file from GCS."""
    safe_filename = secure_filename(filename)

    # Download from GCS to temp file
    temp_path = gcs_storage.download_to_temp(safe_filename, gcs_storage.OUTPUTS_FOLDER)
    if temp_path:
        return send_file(temp_path, as_attachment=True, download_name=safe_filename)
    return jsonify({'error': 'File not found'}), 404


@app.route('/api/files')
@login_required
def get_files():
    """Get list of uploaded files."""
    files = get_uploaded_files()
    # Convert datetime to string for JSON
    for key, value in files.items():
        if value and 'modified' in value:
            value['modified'] = value['modified'].isoformat()
    return jsonify(files)


@app.route('/api/reports')
@login_required
def get_reports():
    """Get list of available reports."""
    reports = get_available_reports()
    # Convert datetime to string for JSON
    for r in reports:
        r['modified'] = r['modified'].isoformat()
    return jsonify(reports)


@app.route('/api/feedback', methods=['POST'])
@login_required
def submit_feedback():
    """Submit user feedback with optional file attachment (screenshot or Excel)."""
    # Support both JSON and multipart form data
    if request.content_type and 'multipart/form-data' in request.content_type:
        message = request.form.get('message', '').strip()
        category = request.form.get('category', '').strip()
        priority = request.form.get('priority', 'Medium')
        page = request.form.get('page', '')
        uploaded_file = request.files.get('file')
    else:
        data = request.get_json()
        if not data:
            return jsonify({'error': 'No data provided'}), 400
        message = data.get('message', '').strip()
        category = data.get('category', '').strip()
        priority = data.get('priority', 'Medium')
        page = data.get('page', '')
        uploaded_file = None

    if not message:
        return jsonify({'error': 'Message is required'}), 400
    if not category:
        return jsonify({'error': 'Category is required'}), 400

    feedback_entry = {
        'username': current_user.username,
        'category': category,
        'priority': priority,
        'page': page,
        'message': message,
        'submitted_at': datetime.now().isoformat(),
        'status': 'New',
        'dev_status': 'unprocessed',
        'attachment': None
    }

    # Handle file upload if present
    if uploaded_file and uploaded_file.filename:
        filename = secure_filename(uploaded_file.filename)
        ext = filename.rsplit('.', 1)[1].lower() if '.' in filename else ''

        allowed_exts = {'png', 'jpg', 'jpeg', 'gif', 'bmp', 'xlsx', 'xls', 'csv'}
        if ext not in allowed_exts:
            return jsonify({'error': f'File type .{ext} not allowed. Accepted: {", ".join(allowed_exts)}'}), 400

        # Check file size (25 MB max)
        uploaded_file.seek(0, 2)
        file_size = uploaded_file.tell()
        uploaded_file.seek(0)
        if file_size > 25 * 1024 * 1024:
            return jsonify({'error': 'File too large. Maximum 25 MB.'}), 400

        # Determine storage folder based on category
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        if category == 'Example File':
            # Example files go to a dedicated dev folder for analysis
            storage_filename = f"example_{timestamp}_{filename}"
            storage_folder = 'feedback/example_files'
        else:
            storage_filename = f"feedback_{timestamp}_{filename}"
            storage_folder = 'feedback/attachments'

        try:
            gcs_storage.upload_file_object(uploaded_file, storage_filename, storage_folder)
            feedback_entry['attachment'] = {
                'filename': filename,
                'stored_as': storage_filename,
                'folder': storage_folder,
                'size': file_size,
                'type': ext,
            }
            print(f"[Feedback] Uploaded attachment: {storage_filename} to {storage_folder}")
        except Exception as e:
            print(f"[ERROR] Failed to upload feedback attachment: {e}")
            return jsonify({'error': 'Failed to upload file'}), 500

    try:
        gcs_storage.save_feedback(feedback_entry)
        return jsonify({'success': True, 'has_attachment': feedback_entry['attachment'] is not None})
    except Exception as e:
        print(f"[ERROR] Failed to save feedback: {e}")
        return jsonify({'error': 'Failed to save feedback'}), 500


@app.route('/api/feedback/mine')
@login_required
def get_my_feedback():
    """Get the current user's own feedback submissions with status."""
    try:
        feedback = gcs_storage.load_feedback()
        my_feedback = [f for f in feedback if f.get('username') == current_user.username]
        my_feedback.reverse()  # Newest first
        return jsonify({'feedback': my_feedback})
    except Exception as e:
        print(f"[ERROR] Failed to load user feedback: {e}")
        return jsonify({'feedback': []})


@app.route('/api/feedback')
@login_required
def get_feedback():
    """Get all feedback (admin only)."""
    if current_user.role != 'admin':
        return jsonify({'error': 'Unauthorized'}), 403

    try:
        feedback = gcs_storage.load_feedback()
        # Return newest first
        feedback.reverse()
        return jsonify({'feedback': feedback})
    except Exception as e:
        print(f"[ERROR] Failed to load feedback: {e}")
        return jsonify({'feedback': []})


@app.route('/api/feedback/export')
@login_required
def export_feedback():
    """Export all feedback as downloadable JSON for dev sessions.
    Includes full feedback entries with attachment metadata.
    Usage: curl -o feedback.json https://dynabot.biz/api/feedback/export
    """
    if current_user.role != 'admin':
        return jsonify({'error': 'Unauthorized'}), 403

    try:
        feedback = gcs_storage.load_feedback()
        feedback.reverse()  # Newest first

        export = {
            'exported_at': datetime.now().isoformat(),
            'total_entries': len(feedback),
            'entries': feedback,
            'attachment_download_url': '/api/feedback/download/{filename}?folder={folder}'
        }

        response = app.response_class(
            response=json.dumps(export, indent=2, default=str),
            status=200,
            mimetype='application/json'
        )
        response.headers['Content-Disposition'] = 'attachment; filename=dynabot_feedback_export.json'
        return response
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/feedback/download/<filename>')
@login_required
def download_feedback_file(filename):
    """Download a feedback attachment from GCS."""
    if current_user.role != 'admin':
        return jsonify({'error': 'Unauthorized'}), 403

    safe_filename = secure_filename(filename)
    folder = request.args.get('folder', 'feedback/attachments')

    # Only allow downloading from feedback folders
    if not folder.startswith('feedback/'):
        return jsonify({'error': 'Invalid folder'}), 400

    temp_path = gcs_storage.download_to_temp(safe_filename, folder)
    if temp_path:
        return send_file(temp_path, as_attachment=True, download_name=safe_filename)
    return jsonify({'error': 'File not found'}), 404


@app.route('/api/feedback/<int:index>/status', methods=['PUT'])
@login_required
def update_feedback_status(index):
    """Update the status of a feedback entry (admin only)."""
    if current_user.role != 'admin':
        return jsonify({'error': 'Unauthorized'}), 403

    data = request.get_json()
    new_status = data.get('status', '').strip() if data else ''

    valid_statuses = ['New', 'In-Work', 'Fixed', 'Resolved w/o Action']
    if new_status not in valid_statuses:
        return jsonify({'error': f'Invalid status. Must be one of: {", ".join(valid_statuses)}'}), 400

    try:
        feedback = gcs_storage.load_feedback()
        # Index is newest-first from the frontend, so reverse to match storage order
        reversed_idx = len(feedback) - 1 - index
        if reversed_idx < 0 or reversed_idx >= len(feedback):
            return jsonify({'error': 'Feedback entry not found'}), 404

        feedback[reversed_idx]['status'] = new_status
        feedback[reversed_idx]['status_updated_by'] = current_user.username
        feedback[reversed_idx]['status_updated_at'] = datetime.now().isoformat()

        # Save the full list back (overwrite)
        if gcs_storage.USE_LOCAL_STORAGE:
            gcs_storage._local_save_json(gcs_storage.FEEDBACK_FILE, feedback)
        else:
            bucket = gcs_storage.get_bucket()
            blob = bucket.blob(gcs_storage.FEEDBACK_FILE)
            blob.upload_from_string(
                json.dumps(feedback, default=str),
                content_type='application/json'
            )

        return jsonify({'success': True, 'status': new_status})
    except Exception as e:
        print(f"[ERROR] Failed to update feedback status: {e}")
        return jsonify({'error': 'Failed to update status'}), 500


@app.route('/api/feedback/<int:index>/dev-status', methods=['PUT'])
@login_required
def update_feedback_dev_status(index):
    """Update the dev pipeline status of a feedback entry (admin only).
    Used by the feedback processing pipeline to track ingestion state.
    """
    if current_user.role != 'admin':
        return jsonify({'error': 'Unauthorized'}), 403

    data = request.get_json()
    new_status = data.get('dev_status', '').strip() if data else ''

    valid_statuses = ['unprocessed', 'ingested', 'actioned', 'closed']
    if new_status not in valid_statuses:
        return jsonify({'error': f'Invalid dev_status. Must be one of: {", ".join(valid_statuses)}'}), 400

    try:
        feedback = gcs_storage.load_feedback()
        reversed_idx = len(feedback) - 1 - index
        if reversed_idx < 0 or reversed_idx >= len(feedback):
            return jsonify({'error': 'Feedback entry not found'}), 404

        feedback[reversed_idx]['dev_status'] = new_status
        feedback[reversed_idx]['dev_status_updated_by'] = current_user.username
        feedback[reversed_idx]['dev_status_updated_at'] = datetime.now().isoformat()

        if gcs_storage.USE_LOCAL_STORAGE:
            gcs_storage._local_save_json(gcs_storage.FEEDBACK_FILE, feedback)
        else:
            bucket = gcs_storage.get_bucket()
            blob = bucket.blob(gcs_storage.FEEDBACK_FILE)
            blob.upload_from_string(
                json.dumps(feedback, default=str),
                content_type='application/json'
            )

        return jsonify({'success': True, 'dev_status': new_status})
    except Exception as e:
        print(f"[ERROR] Failed to update feedback dev_status: {e}")
        return jsonify({'error': 'Failed to update dev_status'}), 500


@app.route('/api/simulation-data')
@login_required
def get_simulation_data():
    """Get simulation data for visual factory floor animation.
    Defaults to the published schedule if available, falls back to current schedule."""

    # Station layout configuration (x, y positions for rendering)
    stations = [
        {'id': 'BLAST', 'name': 'BLAST', 'x': 50, 'y': 180, 'width': 80, 'height': 50},
        {'id': 'TUBE PREP', 'name': 'TUBE PREP', 'x': 180, 'y': 100, 'width': 100, 'height': 50, 'capacity': 18},
        {'id': 'CORE OVEN', 'name': 'CORE OVEN', 'x': 180, 'y': 260, 'width': 100, 'height': 50, 'capacity': 12},
        {'id': 'ASSEMBLY', 'name': 'ASSEMBLY', 'x': 340, 'y': 180, 'width': 80, 'height': 50},
        {'id': 'INJECTION', 'name': 'INJECTION', 'x': 470, 'y': 180, 'width': 100, 'height': 80, 'machines': ['D1', 'D2', 'D3', 'D4', 'D5']},
        {'id': 'CURE', 'name': 'CURE', 'x': 620, 'y': 180, 'width': 80, 'height': 50, 'capacity': 16},
        {'id': 'QUENCH', 'name': 'QUENCH', 'x': 620, 'y': 280, 'width': 80, 'height': 50, 'capacity': 16},
        {'id': 'DISASSEMBLY', 'name': 'DISASSEMBLY', 'x': 470, 'y': 330, 'width': 100, 'height': 50},
        {'id': 'BLD END CUTBACK', 'name': 'CUTBACK', 'x': 340, 'y': 330, 'width': 80, 'height': 50},
        {'id': 'INJ END CUTBACK', 'name': 'CUTBACK', 'x': 340, 'y': 330, 'width': 80, 'height': 50},
        {'id': 'CUT THREADS', 'name': 'CUT THREADS', 'x': 210, 'y': 330, 'width': 80, 'height': 50},
        {'id': 'INSPECT', 'name': 'INSPECT', 'x': 80, 'y': 330, 'width': 80, 'height': 50},
    ]

    # Try in-memory objects first (freshly generated)
    if current_schedule['orders']:
        all_starts = []
        all_ends = []
        for order in current_schedule['orders']:
            if order.blast_date:
                all_starts.append(order.blast_date)
            if order.completion_date:
                all_ends.append(order.completion_date)
            for op in order.operations:
                all_starts.append(op.start_time)
                all_ends.append(op.end_time)

        start_date = min(all_starts) if all_starts else datetime.now()
        end_date = max(all_ends) if all_ends else datetime.now()

        parts = []
        orders_with_ops = 0
        for order in current_schedule['orders']:
            operations = []
            for op in order.operations:
                operations.append({
                    'station': op.operation_name,
                    'start': op.start_time.isoformat(),
                    'end': op.end_time.isoformat(),
                    'resource': op.resource_id
                })
            if operations:
                orders_with_ops += 1
            parts.append({
                'wo_number': order.wo_number or '',
                'part_number': order.part_number or '',
                'customer': order.customer or '',
                'priority': order.priority or 'Normal',
                'rubber_type': order.rubber_type or '',
                'assigned_core': order.assigned_core or '',
                'is_rework': order.is_reline,
                'operations': operations
            })

        print(f"[Simulation API] {len(parts)} parts, {orders_with_ops} with operations (from in-memory)")

        # Save for future use (so simulation survives server restarts)
        sim_payload = {
            'schedule_info': {
                'start_date': start_date.isoformat(),
                'end_date': end_date.isoformat(),
                'total_orders': len(parts),
                'generated_at': current_schedule['generated_at'].isoformat() if current_schedule.get('generated_at') else None
            },
            'stations': stations,
            'parts': parts
        }
        gcs_storage.save_simulation_data(sim_payload)

        return jsonify(sim_payload)

    # Fall back to persisted simulation data (published schedule)
    persisted_sim = gcs_storage.load_simulation_data()
    if persisted_sim:
        print(f"[Simulation API] Serving from persisted simulation data (published schedule)")
        # Ensure stations are current (layout may have been updated)
        persisted_sim['stations'] = stations
        return jsonify(persisted_sim)

    return jsonify({'error': 'No schedule available. Please generate or publish a schedule first.'}), 400


# ============== Planner Workflow API ==============


# Scenario label map for display
SCENARIO_CONFIGS = {
    '4day_10h': {'working_days': [0, 1, 2, 3], 'shift_hours': 10, 'label': '4 Days x 10 Hours'},
    '4day_12h': {'working_days': [0, 1, 2, 3], 'shift_hours': 12, 'label': '4 Days x 12 Hours'},
    '5day_12h': {'working_days': [0, 1, 2, 3, 4], 'shift_hours': 12, 'label': '5 Days x 12 Hours'},
    '6day_12h': {'working_days': [0, 1, 2, 3, 4, 5], 'shift_hours': 12, 'label': '6 Days x 12 Hours'},
}


def _compute_stats_from_serialized(serialized_orders):
    """Compute stats dict from a list of serialized order dicts."""
    AT_RISK_BUFFER_DAYS = 2
    total = len(serialized_orders)
    on_time = sum(1 for o in serialized_orders if o.get('on_time_status') == 'On Time')
    late = sum(1 for o in serialized_orders if o.get('on_time_status') == 'Late')
    at_risk = sum(1 for o in serialized_orders if o.get('on_time_status') == 'At Risk')
    turnarounds = [o['turnaround_days'] for o in serialized_orders if o.get('turnaround_days')]
    avg_turnaround = round(sum(turnarounds) / len(turnarounds), 1) if turnarounds else 0
    return {
        'total_orders': total,
        'on_time': on_time,
        'late': late,
        'at_risk': at_risk,
        'avg_turnaround': avg_turnaround,
    }


@app.route('/api/planner/simulate-scenarios', methods=['POST'])
@login_required
def simulate_scenarios():
    """
    Step 2: Simulate base schedule across standard scenarios.
    Runs the DES engine for each preset (4dx10h, 4dx12h, 5dx12h, 6dx12h) WITHOUT hot list.
    Returns comparison metrics for each scenario.
    """
    global planner_state

    if current_user.role not in ('admin', 'planner'):
        return jsonify({'error': 'Only Planner and Admin users can simulate schedules.'}), 403

    try:
        import tempfile
        import shutil

        temp_dir = tempfile.mkdtemp(prefix='estradabot_scenarios_')
        print(f"[Planner] Downloading files from GCS to {temp_dir}")

        local_paths = gcs_storage.download_files_for_processing(temp_dir)
        loader = DataLoader(data_dir=temp_dir)
        loader.load_all()

        if not loader.orders:
            shutil.rmtree(temp_dir, ignore_errors=True)
            return jsonify({'error': 'No orders loaded. Please upload a Sales Order file.'}), 400

        # Exclude orders on hold
        order_holds = gcs_storage.load_order_holds()
        if order_holds:
            before_count = len(loader.orders)
            loader.orders = [o for o in loader.orders if o.get('wo_number') not in order_holds]
            held_count = before_count - len(loader.orders)
            if held_count > 0:
                print(f"[Planner] Excluded {held_count} orders on hold")

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        scenarios = {}

        for scenario_key, config in SCENARIO_CONFIGS.items():
            mode_label = scenario_key.replace('_', '_').upper()
            print(f"[Planner] Running scenario: {config['label']}...")

            result = _run_schedule_mode(
                loader,
                config['working_days'],
                mode_label,
                temp_dir,
                timestamp,
                shift_hours=config['shift_hours'],
                skip_hot_list=True  # Base schedule = no hot list/special requests
            )

            scenarios[scenario_key] = {
                'label': config['label'],
                'stats': result['stats'],
                'serialized_orders': result['serialized_orders'],
                'orders': result['orders'],
                'baseline_orders': result['baseline_orders'],
                'reports': result['reports'],
            }

        # Store in planner state (keep loader for later re-runs with requests)
        planner_state['scenarios'] = scenarios
        planner_state['simulated_at'] = datetime.now()
        planner_state['loader'] = loader
        planner_state['base_scenario'] = None
        planner_state['base_schedule'] = None
        planner_state['final_schedule'] = None
        planner_state['_temp_dir'] = temp_dir  # Keep for later report generation

        # Return comparison metrics
        comparison = {}
        for key, scenario in scenarios.items():
            comparison[key] = {
                'label': scenario['label'],
                'stats': scenario['stats'],
            }

        return jsonify({
            'success': True,
            'scenarios': comparison,
            'simulated_at': planner_state['simulated_at'].isoformat()
        })

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/api/planner/set-base-schedule', methods=['POST'])
@login_required
def set_base_schedule():
    """
    Step 3: Planner selects one of the scenarios as the base schedule.
    Expects JSON body: { "scenario": "4day_10h" | "4day_12h" | "5day_12h" | "6day_12h" | "custom" }
    """
    global planner_state

    if current_user.role not in ('admin', 'planner'):
        return jsonify({'error': 'Unauthorized'}), 403

    data = request.get_json()
    scenario_key = data.get('scenario')

    # Accept standard presets and the custom scenario
    if scenario_key not in SCENARIO_CONFIGS and scenario_key != 'custom':
        return jsonify({'error': f'Invalid scenario: {scenario_key}'}), 400

    if not planner_state.get('scenarios'):
        return jsonify({'error': 'No scenarios simulated. Run "Simulate Base Schedule" first.'}), 400

    scenario = planner_state['scenarios'].get(scenario_key)
    if not scenario:
        return jsonify({'error': f'Scenario {scenario_key} not found in results.'}), 400

    planner_state['base_scenario'] = scenario_key
    planner_state['base_schedule'] = scenario

    return jsonify({
        'success': True,
        'selected': scenario_key,
        'label': scenario['label'],
        'stats': scenario['stats']
    })


@app.route('/api/planner/simulate-custom-scenario', methods=['POST'])
@login_required
def simulate_custom_scenario():
    """
    Run a single custom simulation with per-day shift configuration.
    Accepts a Mon-Sat grid of per-day configs (full/skeleton, day/night/both, takt).
    Returns one scenario result that can be selected as the base schedule.

    Expected JSON body:
    {
        "shift_hours": 12,
        "days": {
            "0": {"working": true, "shift_mode": "full", "active_shifts": "both", "takt": 30},
            "1": {"working": true, "shift_mode": "skeleton", "active_shifts": "day", "takt": 60},
            ...
        }
    }
    """
    global planner_state

    if current_user.role not in ('admin', 'planner'):
        return jsonify({'error': 'Only Planner and Admin users can simulate schedules.'}), 403

    data = request.get_json()
    if not data or 'days' not in data:
        return jsonify({'error': 'Missing day configuration.'}), 400

    shift_hours = data.get('shift_hours', 12)
    days_raw = data['days']

    # Validate and build config
    working_days = []
    day_configs = {}
    errors = []

    for day_str, day_data in days_raw.items():
        try:
            weekday = int(day_str)
        except (ValueError, TypeError):
            errors.append(f'Invalid day key: {day_str}')
            continue

        if weekday < 0 or weekday > 5:
            errors.append(f'Day must be 0-5 (Mon-Sat), got {weekday}')
            continue

        if not day_data.get('working', False):
            continue

        working_days.append(weekday)

        takt = day_data.get('takt', 30)
        if not isinstance(takt, (int, float)) or takt < 1 or takt > 120:
            errors.append(f'Takt must be 1-120 minutes, got {takt} for day {weekday}')
            continue

        active_shifts = day_data.get('active_shifts', 'both')
        shift_mode = day_data.get('shift_mode', 'full')

        # Validate: 10h base disables night-only skeleton
        if shift_hours == 10 and active_shifts == 'night':
            errors.append(f'Night-only is not available with 10-hour shifts (day {weekday})')
            continue

        # Every working day needs at least one shift
        if active_shifts not in ('day', 'night', 'both'):
            errors.append(f'Invalid active_shifts: {active_shifts} for day {weekday}')
            continue

        day_configs[weekday] = DayShiftConfig(
            shift_mode=shift_mode,
            active_shifts=active_shifts,
            takt_time_minutes=int(takt)
        )

    if not working_days:
        errors.append('At least one working day is required')

    if errors:
        return jsonify({'error': 'Validation errors: ' + '; '.join(errors)}), 400

    working_days.sort()

    # Build a descriptive label
    day_names = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat']
    skeleton_days = [day_names[d] for d in working_days if day_configs.get(d, DayShiftConfig()).shift_mode == 'skeleton']
    label = f'Custom ({len(working_days)}d x {shift_hours}h'
    if skeleton_days:
        label += f', skeleton: {",".join(skeleton_days)}'
    label += ')'

    try:
        import tempfile
        import shutil

        # Reuse existing temp_dir from scenario simulation if available
        temp_dir = planner_state.get('_temp_dir')
        loader = planner_state.get('loader')

        if not temp_dir or not loader:
            # Need to load fresh data
            temp_dir = tempfile.mkdtemp(prefix='estradabot_custom_')
            gcs_storage.download_files_for_processing(temp_dir)
            loader = DataLoader(data_dir=temp_dir)
            loader.load_all()

            if not loader.orders:
                shutil.rmtree(temp_dir, ignore_errors=True)
                return jsonify({'error': 'No orders loaded. Please upload a Sales Order file.'}), 400

            # Exclude orders on hold
            order_holds = gcs_storage.load_order_holds()
            if order_holds:
                loader.orders = [o for o in loader.orders if o.get('wo_number') not in order_holds]

            planner_state['loader'] = loader
            planner_state['_temp_dir'] = temp_dir

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        mode_label = 'CUSTOM'
        print(f"[Planner] Running custom scenario: {label}...")

        result = _run_schedule_mode(
            loader,
            working_days,
            mode_label,
            temp_dir,
            timestamp,
            shift_hours=shift_hours,
            skip_hot_list=True,
            day_configs=day_configs
        )

        scenario_data = {
            'label': label,
            'stats': result['stats'],
            'serialized_orders': result['serialized_orders'],
            'orders': result['orders'],
            'baseline_orders': result['baseline_orders'],
            'reports': result['reports'],
        }

        # Store in planner state alongside standard scenarios
        if 'scenarios' not in planner_state or planner_state['scenarios'] is None:
            planner_state['scenarios'] = {}
        planner_state['scenarios']['custom'] = scenario_data

        return jsonify({
            'success': True,
            'scenario': {
                'label': label,
                'stats': result['stats'],
            }
        })

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


REDLINE_TYPE_LABELS = {
    'rubber_override': 'Rubber Substitution',
    'cutback_length': 'Cutback Length Change',
    'hip_cap_injection': 'Hip & Cap Injection',
    'full_restrict_injection': 'Full Restrict Injection',
}

def _build_special_instructions(req):
    """Build special instructions string from an approved request.
    Format: "Request Type" - "Rubber Override" (if applicable) - "Reason/Comments"
    """
    request_type = req.get('request_type', '')
    label = REDLINE_TYPE_LABELS.get(request_type)
    if not label:
        return ''  # Not a redline type — no special instructions

    parts = [label]
    rubber = req.get('rubber_override')
    if rubber:
        parts.append(rubber)
    reason = (req.get('reason') or req.get('comments') or '').strip()
    if reason:
        parts.append(reason)
    return ' - '.join(parts)


@app.route('/api/special-requests', methods=['GET'])
@login_required
def get_special_requests():
    """Get all special requests, optionally filtered by status.
    Auto-expires unmatched Mode B placeholders older than 28 days.
    Auto-purges any pending request older than 14 days."""
    status_filter = request.args.get('status')
    requests_list = gcs_storage.load_special_requests()

    now = datetime.now()
    expired_count = 0
    for req in requests_list:
        if req.get('status') != 'pending' or not req.get('submitted_at'):
            continue
        try:
            submitted = datetime.fromisoformat(req['submitted_at'])
            age_days = (now - submitted).days
            if req.get('matched') is False:
                # Unmatched Mode B placeholders get a 28-day grace period
                if age_days >= 28:
                    req['status'] = 'expired'
                    req['expired_at'] = now.isoformat()
                    req['expiry_reason'] = 'Unmatched placeholder expired after 28 days'
                    expired_count += 1
            elif age_days >= 14:
                # All other pending requests: purge after 14 days
                req['status'] = 'expired'
                req['expired_at'] = now.isoformat()
                req['expiry_reason'] = 'Pending request not processed within 14 days'
                expired_count += 1
        except (ValueError, TypeError):
            pass

    if expired_count > 0:
        gcs_storage.save_special_requests(requests_list)
        print(f"[SpecialRequests] Expired {expired_count} stale pending request(s)")

    if status_filter:
        requests_list = [r for r in requests_list if r.get('status') == status_filter]

    return jsonify({'requests': requests_list})


@app.route('/api/special-requests', methods=['POST'])
@login_required
def create_special_request():
    """
    Submit a new special request (hot list entry from the app).
    Available to customer_service, planner, admin roles.
    """
    allowed_roles = ('admin', 'planner', 'customer_service')
    if current_user.role not in allowed_roles:
        return jsonify({'error': 'Your role cannot submit special requests.'}), 403

    data = request.get_json()
    if not data:
        return jsonify({'error': 'No data provided'}), 400

    wo_number = data.get('wo_number', '').strip()
    if not wo_number:
        return jsonify({'error': 'Work Order number is required'}), 400

    request_entry = {
        'id': f"SR-{datetime.now().strftime('%Y%m%d%H%M%S')}-{wo_number}",
        'wo_number': wo_number,
        'request_type': data.get('request_type', 'hot_list'),  # hot_list, expedite, rubber_override
        'is_asap': data.get('is_asap', False),
        'need_by_date': data.get('need_by_date'),
        'rubber_override': data.get('rubber_override'),
        'reason': data.get('reason', ''),
        'comments': data.get('comments', ''),
        'submitted_by': current_user.username,
        'submitted_at': datetime.now().isoformat(),
        'status': 'pending',  # pending, approved, rejected, published
        'reviewed_by': None,
        'reviewed_at': None,
        'rejection_reason': None,
        'matched': data.get('matched', True),  # False = Mode B placeholder (WO not in system)
    }

    # Load existing and append
    all_requests = gcs_storage.load_special_requests()
    all_requests.append(request_entry)
    gcs_storage.save_special_requests(all_requests)

    return jsonify({
        'success': True,
        'request': request_entry
    })


@app.route('/api/order-holds', methods=['GET'])
@login_required
def get_order_holds():
    """Get all current order holds."""
    holds = gcs_storage.load_order_holds()
    return jsonify({'holds': holds})


@app.route('/api/order-holds', methods=['POST'])
@login_required
def set_order_hold():
    """Place an order on hold. Requires admin or planner role."""
    if current_user.role not in ('admin', 'planner'):
        return jsonify({'error': 'Only Planner and Admin users can place holds.'}), 403

    data = request.get_json()
    wo_number = data.get('wo_number', '').strip()
    reason = data.get('reason', '').strip()

    if not wo_number:
        return jsonify({'error': 'Work Order number is required'}), 400

    holds = gcs_storage.load_order_holds()
    holds[wo_number] = {
        'held_by': current_user.username,
        'held_at': datetime.now().isoformat(),
        'reason': reason
    }
    gcs_storage.save_order_holds(holds)

    return jsonify({'success': True, 'wo_number': wo_number, 'hold': holds[wo_number]})


@app.route('/api/order-holds/<wo_number>', methods=['DELETE'])
@login_required
def remove_order_hold(wo_number):
    """Remove a hold from an order. Requires admin or planner role."""
    if current_user.role not in ('admin', 'planner'):
        return jsonify({'error': 'Only Planner and Admin users can remove holds.'}), 403

    holds = gcs_storage.load_order_holds()
    if wo_number in holds:
        del holds[wo_number]
        gcs_storage.save_order_holds(holds)
        return jsonify({'success': True, 'wo_number': wo_number})
    else:
        return jsonify({'error': f'No hold found for WO {wo_number}'}), 404


@app.route('/api/special-requests/impact-preview', methods=['POST'])
@login_required
def impact_preview():
    """
    Preview the scheduling impact of a single request before submitting.
    Uses the published schedule (or current schedule) as baseline,
    then re-runs DES with the proposed request added.
    Returns impact data showing which orders are affected.
    """
    data = request.get_json()
    if not data:
        return jsonify({'error': 'No data provided'}), 400

    wo_number = data.get('wo_number', '').strip()
    if not wo_number:
        return jsonify({'error': 'Work Order number is required'}), 400

    # Check if we have a published or current schedule to use as baseline
    pub = gcs_storage.load_published_schedule()
    has_published = pub is not None

    # We need the loader and a scenario config to re-run the scheduler
    # Use the planner_state loader if available, otherwise try to create one
    loader = planner_state.get('loader')
    if not loader:
        # Try to create a loader from uploaded files
        import tempfile
        temp_dir = tempfile.mkdtemp(prefix='estradabot_preview_')
        try:
            local_paths = gcs_storage.download_files_for_processing(temp_dir)
            loader = DataLoader(local_paths)
            if not loader.orders:
                return jsonify({'error': 'No schedule data available. Upload files and generate a schedule first.'}), 400
        except Exception as e:
            return jsonify({'error': f'Could not load data for simulation: {str(e)}'}), 500

    # Determine which scenario config to use
    scenario_key = planner_state.get('base_scenario')
    if not scenario_key and pub:
        scenario_key = pub.get('scenario_key')
    if not scenario_key:
        scenario_key = '4day_12h'  # Sensible default

    config = SCENARIO_CONFIGS.get(scenario_key, SCENARIO_CONFIGS['4day_12h'])

    try:
        # Run baseline (without this request)
        scheduler_baseline = DESScheduler(
            orders=loader.orders,
            core_mapping=loader.core_mapping,
            core_inventory=loader.core_inventory,
            working_days=config['working_days'],
            shift_hours=config['shift_hours'],
            wip_orders=loader.wip_in_process_orders
        )
        baseline_orders = scheduler_baseline.schedule_orders()

        # Build a single-entry hot list for the proposed request
        preview_hot_list = [{
            'wo_number': wo_number,
            'is_asap': data.get('is_asap', False),
            'need_by_date': data.get('need_by_date'),
            'date_req_made': datetime.now().isoformat(),
            'rubber_override': data.get('rubber_override'),
            'row_position': 9999,
            'comments': data.get('reason', ''),
            'core': '',
            'item': '',
            'description': '',
            'customer': '',
            'source': 'impact_preview',
            'special_instructions': _build_special_instructions(data),
        }]

        # Run with the proposed request
        scheduler_with = DESScheduler(
            orders=loader.orders,
            core_mapping=loader.core_mapping,
            core_inventory=loader.core_inventory,
            working_days=config['working_days'],
            shift_hours=config['shift_hours'],
            wip_orders=loader.wip_in_process_orders
        )
        orders_with = scheduler_with.schedule_orders(hot_list_entries=preview_hot_list)

        # Build impact analysis
        baseline_lookup = {o.wo_number: o for o in baseline_orders}
        impact_items = []
        total_delay_hours = 0
        orders_now_late = 0

        for order in orders_with:
            if order.wo_number == wo_number:
                continue

            baseline = baseline_lookup.get(order.wo_number)
            if not baseline or not baseline.blast_date or not order.blast_date:
                continue

            delay_hours = (order.blast_date - baseline.blast_date).total_seconds() / 3600
            if delay_hours > 0.5:
                was_on_time = baseline.on_time
                is_on_time = order.on_time
                status_change = ''
                if was_on_time and not is_on_time:
                    status_change = 'NOW LATE'
                    orders_now_late += 1

                impact_items.append({
                    'wo_number': order.wo_number,
                    'part_number': order.part_number,
                    'customer': order.customer,
                    'delay_hours': round(delay_hours, 1),
                    'status_change': status_change,
                })
                total_delay_hours += delay_hours

        impact_items.sort(key=lambda x: -x['delay_hours'])

        return jsonify({
            'success': True,
            'has_published_schedule': has_published,
            'scenario': scenario_key,
            'impact': {
                'total_delayed': len(impact_items),
                'total_delay_hours': round(total_delay_hours, 1),
                'orders_now_late': orders_now_late,
                'items': impact_items[:20],
            },
        })

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'Impact simulation failed: {str(e)}'}), 500


@app.route('/api/planner/simulate-with-requests', methods=['POST'])
@login_required
def simulate_with_requests():
    """
    Step 6: Simulate the base schedule with hot list + approved special requests applied.
    Returns impact analysis showing how each request affects the schedule.
    """
    global planner_state

    if current_user.role not in ('admin', 'planner'):
        return jsonify({'error': 'Unauthorized'}), 403

    if not planner_state.get('base_scenario'):
        return jsonify({'error': 'No base schedule selected. Set a base scenario first.'}), 400

    loader = planner_state.get('loader')
    if not loader:
        return jsonify({'error': 'Data loader not available. Re-run "Simulate Base Schedule".'}), 400

    try:
        scenario_key = planner_state['base_scenario']
        config = SCENARIO_CONFIGS[scenario_key]

        # Gather approved special requests and combine with hot list entries
        all_requests = gcs_storage.load_special_requests()
        pending_requests = [r for r in all_requests if r.get('status') == 'pending']

        # Convert special requests to hot list entry format for the scheduler
        combined_hot_list = list(loader.hot_list_entries) if loader.hot_list_entries else []

        for req in pending_requests:
            combined_hot_list.append({
                'wo_number': req['wo_number'],
                'is_asap': req.get('is_asap', False),
                'need_by_date': req.get('need_by_date'),
                'date_req_made': req.get('submitted_at'),
                'rubber_override': req.get('rubber_override'),
                'row_position': 9999,  # App requests after file-based entries
                'comments': req.get('comments', ''),
                'core': '',
                'item': '',
                'description': '',
                'customer': '',
                'source': 'app_request',
                'request_id': req['id'],
            })

        # Run scheduler with combined hot list on the base scenario config
        scheduler_with_requests = DESScheduler(
            orders=loader.orders,
            core_mapping=loader.core_mapping,
            core_inventory=loader.core_inventory,
            working_days=config['working_days'],
            shift_hours=config['shift_hours'],
            wip_orders=loader.wip_in_process_orders
        )
        scheduled_with_requests = scheduler_with_requests.schedule_orders(
            hot_list_entries=combined_hot_list
        )

        # Get baseline orders (without requests) from the stored base schedule
        baseline_orders = planner_state['base_schedule']['orders']

        # Build impact analysis data
        baseline_lookup = {o.wo_number: o for o in baseline_orders}
        hot_list_wos = {e['wo_number'] for e in combined_hot_list}

        impact_items = []
        total_delay_hours = 0
        orders_now_late = 0

        for order in scheduled_with_requests:
            if order.wo_number in hot_list_wos:
                continue  # Skip the hot list orders themselves

            baseline = baseline_lookup.get(order.wo_number)
            if not baseline or not baseline.blast_date or not order.blast_date:
                continue

            delay_hours = (order.blast_date - baseline.blast_date).total_seconds() / 3600
            if delay_hours > 0.5:
                was_on_time = baseline.on_time
                is_on_time = order.on_time
                status_change = ''
                if was_on_time and not is_on_time:
                    status_change = 'NOW LATE'
                    orders_now_late += 1

                impact_items.append({
                    'wo_number': order.wo_number,
                    'part_number': order.part_number,
                    'customer': order.customer,
                    'delay_hours': round(delay_hours, 1),
                    'was_on_time': was_on_time,
                    'is_on_time': is_on_time,
                    'status_change': status_change,
                })
                total_delay_hours += delay_hours

        impact_items.sort(key=lambda x: -x['delay_hours'])

        # Serialize request-applied orders for the final schedule
        AT_RISK_BUFFER_DAYS = 2
        serialized = []
        for order in scheduled_with_requests:
            deadline = order.basic_finish_date or order.promise_date
            if not order.on_time:
                status = 'Late'
            elif deadline and order.completion_date:
                days_buffer = (deadline - order.completion_date).days
                status = 'At Risk' if days_buffer <= AT_RISK_BUFFER_DAYS else 'On Time'
            else:
                status = 'On Time'

            serialized.append({
                'wo_number': order.wo_number or '',
                'serial_number': order.serial_number or '',
                'part_number': order.part_number or '',
                'description': order.description or '',
                'customer': order.customer or '',
                'assigned_core': order.assigned_core or '',
                'rubber_type': order.rubber_type or '',
                'priority': order.priority,
                'blast_date': order.blast_date.isoformat() if order.blast_date else None,
                'completion_date': order.completion_date.isoformat() if order.completion_date else None,
                'promise_date': order.promise_date.isoformat() if order.promise_date else None,
                'basic_finish_date': order.basic_finish_date.isoformat() if order.basic_finish_date else None,
                'turnaround_days': order.turnaround_days,
                'on_time': order.on_time,
                'on_time_status': status,
                'is_reline': order.is_reline,
                'special_instructions': order.special_instructions or '',
            })

        stats = _compute_stats_from_serialized(serialized)
        stats['hot_list_count'] = len(combined_hot_list)

        # Store for final schedule generation
        planner_state['_impact_orders'] = scheduled_with_requests
        planner_state['_impact_serialized'] = serialized
        planner_state['_impact_stats'] = stats
        planner_state['_combined_hot_list'] = combined_hot_list

        return jsonify({
            'success': True,
            'impact': {
                'total_delayed': len(impact_items),
                'total_delay_hours': round(total_delay_hours, 1),
                'orders_now_late': orders_now_late,
                'items': impact_items[:50],  # Top 50 most impacted
            },
            'stats_with_requests': stats,
            'hot_list_count': len(combined_hot_list),
            'file_hot_list_count': len(loader.hot_list_entries) if loader.hot_list_entries else 0,
            'app_request_count': len(pending_requests),
        })

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/api/planner/file-hot-list', methods=['GET'])
@login_required
def get_file_hot_list():
    """Return the file-based hot list entries from the current planner session."""
    if current_user.role not in ('admin', 'planner'):
        return jsonify({'error': 'Unauthorized'}), 403

    loader = planner_state.get('loader')
    if not loader or not loader.hot_list_entries:
        return jsonify({'entries': []})

    entries = []
    for e in loader.hot_list_entries:
        need_by = e.get('need_by_date')
        entries.append({
            'wo_number': e.get('wo_number', ''),
            'is_asap': e.get('is_asap', False),
            'need_by_date': need_by.strftime('%Y-%m-%d') if need_by else None,
            'rubber_override': e.get('rubber_override'),
            'comments': e.get('comments', ''),
            'customer': e.get('customer', ''),
            'description': e.get('description', ''),
        })
    return jsonify({'entries': entries})


@app.route('/api/planner/approve-requests', methods=['POST'])
@login_required
def approve_requests():
    """
    Step 7a: Planner approves or rejects individual special requests.
    Expects JSON body: { "approvals": { "request_id": "approved"|"rejected", ... } }
    """
    if current_user.role not in ('admin', 'planner'):
        return jsonify({'error': 'Unauthorized'}), 403

    data = request.get_json()
    approvals = data.get('approvals', {})
    rejection_reason = data.get('rejection_reason', '')

    if not approvals:
        return jsonify({'error': 'No approval decisions provided'}), 400

    all_requests = gcs_storage.load_special_requests()
    updated = 0

    for req in all_requests:
        req_id = req.get('id')
        if req_id in approvals:
            new_status = approvals[req_id]
            if new_status in ('approved', 'rejected'):
                req['status'] = new_status
                req['reviewed_by'] = current_user.username
                req['reviewed_at'] = datetime.now().isoformat()
                if new_status == 'rejected' and rejection_reason:
                    req['rejection_reason'] = rejection_reason
                updated += 1

    gcs_storage.save_special_requests(all_requests)

    return jsonify({
        'success': True,
        'updated': updated
    })


@app.route('/api/planner/generate-final', methods=['POST'])
@login_required
def generate_final_schedule():
    """
    Step 7b: Generate the final schedule with approved requests.
    Re-runs the scheduler with only approved requests included.
    """
    global planner_state

    if current_user.role not in ('admin', 'planner'):
        return jsonify({'error': 'Unauthorized'}), 403

    if not planner_state.get('base_scenario'):
        return jsonify({'error': 'No base schedule set.'}), 400

    loader = planner_state.get('loader')
    if not loader:
        return jsonify({'error': 'Data loader expired. Re-run scenario simulation.'}), 400

    try:
        scenario_key = planner_state['base_scenario']
        config = SCENARIO_CONFIGS[scenario_key]

        # Build hot list from file entries + approved app requests only
        all_requests = gcs_storage.load_special_requests()
        approved_requests = [r for r in all_requests if r.get('status') == 'approved']

        # Planner can optionally restrict which file hot list entries are included.
        # If included_file_wos is provided, only include those WO numbers.
        # If absent or None, include all file entries (default behaviour).
        data = request.get_json(silent=True) or {}
        included_file_wos = data.get('included_file_wos')  # list of WO strings, or None

        file_entries = loader.hot_list_entries if loader.hot_list_entries else []
        if included_file_wos is not None:
            included_set = set(str(w) for w in included_file_wos)
            file_entries = [e for e in file_entries if str(e.get('wo_number', '')) in included_set]

        combined_hot_list = list(file_entries)
        for req in approved_requests:
            combined_hot_list.append({
                'wo_number': req['wo_number'],
                'is_asap': req.get('is_asap', False),
                'need_by_date': req.get('need_by_date'),
                'date_req_made': req.get('submitted_at'),
                'rubber_override': req.get('rubber_override'),
                'row_position': 9999,
                'comments': req.get('comments', ''),
                'core': '',
                'item': '',
                'description': '',
                'customer': '',
                'source': 'app_request',
                'request_id': req['id'],
                'special_instructions': _build_special_instructions(req),
            })

        # Run final schedule
        import tempfile
        temp_dir = planner_state.get('_temp_dir', tempfile.mkdtemp(prefix='estradabot_final_'))
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')

        # Run baseline (for impact analysis reports)
        scheduler_baseline = DESScheduler(
            orders=loader.orders,
            core_mapping=loader.core_mapping,
            core_inventory=loader.core_inventory,
            working_days=config['working_days'],
            shift_hours=config['shift_hours'],
            wip_orders=loader.wip_in_process_orders
        )
        baseline_orders = scheduler_baseline.schedule_orders()

        # Run with hot list
        scheduler_final = DESScheduler(
            orders=loader.orders,
            core_mapping=loader.core_mapping,
            core_inventory=loader.core_inventory,
            working_days=config['working_days'],
            shift_hours=config['shift_hours'],
            wip_orders=loader.wip_in_process_orders
        )
        final_orders = scheduler_final.schedule_orders(hot_list_entries=combined_hot_list)

        # Export reports
        reports = {}
        mode_label = f'Final_{scenario_key}'

        master_filename = f'Master_Schedule_{mode_label}_{timestamp}.xlsx'
        master_path = os.path.join(temp_dir, master_filename)
        from exporters.excel_exporter import export_master_schedule, export_blast_schedule, export_core_schedule, export_pending_core_report
        export_master_schedule(final_orders, master_path)
        gcs_storage.upload_file(master_path, master_filename, gcs_storage.OUTPUTS_FOLDER)
        reports['master'] = master_filename

        blast_filename = f'BLAST_Schedule_{mode_label}_{timestamp}.xlsx'
        blast_path = os.path.join(temp_dir, blast_filename)
        on_blaster = [o for o in loader.wip_in_process_orders if o.get('current_operation') == 1300]
        export_blast_schedule(final_orders, blast_path, currently_blasting=on_blaster)
        gcs_storage.upload_file(blast_path, blast_filename, gcs_storage.OUTPUTS_FOLDER)
        reports['blast'] = blast_filename

        utilization_filename = f'Resource_Utilization_{mode_label}_{timestamp}.xlsx'
        utilization_path = os.path.join(temp_dir, utilization_filename)
        export_resource_utilization(final_orders, utilization_path)
        gcs_storage.upload_file(utilization_path, utilization_filename, gcs_storage.OUTPUTS_FOLDER)
        reports['utilization'] = utilization_filename

        # Impact analysis
        if combined_hot_list:
            hot_list_core_shortages = getattr(scheduler_final, 'hot_list_core_shortages', [])
            impact_path = generate_impact_analysis(
                final_orders, baseline_orders, combined_hot_list,
                hot_list_core_shortages, temp_dir
            )
            impact_filename = os.path.basename(impact_path)
            gcs_storage.upload_file(impact_path, impact_filename, gcs_storage.OUTPUTS_FOLDER)
            reports['impact'] = impact_filename

        # Serialize final orders
        AT_RISK_BUFFER_DAYS = 2
        serialized = []
        on_time_count = late_count = at_risk_count = 0
        for order in final_orders:
            deadline = order.basic_finish_date or order.promise_date
            if not order.on_time:
                status = 'Late'
                late_count += 1
            elif deadline and order.completion_date:
                days_buffer = (deadline - order.completion_date).days
                if days_buffer <= AT_RISK_BUFFER_DAYS:
                    status = 'At Risk'
                    at_risk_count += 1
                else:
                    status = 'On Time'
                    on_time_count += 1
            else:
                status = 'On Time'
                on_time_count += 1

            serialized.append({
                'wo_number': order.wo_number or '',
                'serial_number': order.serial_number or '',
                'part_number': order.part_number or '',
                'description': order.description or '',
                'customer': order.customer or '',
                'assigned_core': order.assigned_core or '',
                'rubber_type': order.rubber_type or '',
                'priority': order.priority,
                'blast_date': order.blast_date.isoformat() if order.blast_date else None,
                'completion_date': order.completion_date.isoformat() if order.completion_date else None,
                'promise_date': order.promise_date.isoformat() if order.promise_date else None,
                'basic_finish_date': order.basic_finish_date.isoformat() if order.basic_finish_date else None,
                'turnaround_days': order.turnaround_days,
                'on_time': order.on_time,
                'on_time_status': status,
                'is_reline': order.is_reline,
                'special_instructions': order.special_instructions or '',
                'supermarket_location': order.supermarket_location or '',
            })

        turnaround_times = [o.turnaround_days for o in final_orders if o.turnaround_days]
        avg_turnaround = sum(turnaround_times) / len(turnaround_times) if turnaround_times else 0

        stats = {
            'total_orders': len(final_orders),
            'on_time': on_time_count,
            'late': late_count,
            'at_risk': at_risk_count,
            'avg_turnaround': round(avg_turnaround, 1),
            'hot_list_count': len(combined_hot_list),
        }

        planner_state['final_schedule'] = {
            'orders': final_orders,
            'serialized_orders': serialized,
            'stats': stats,
            'reports': reports,
            'scenario_key': scenario_key,
            'scenario_label': config['label'],
            'generated_at': datetime.now().isoformat(),
            'approved_request_count': len(approved_requests),
        }

        return jsonify({
            'success': True,
            'stats': stats,
            'reports': reports,
            'scenario_label': config['label'],
            'approved_requests': len(approved_requests),
        })

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/api/planner/publish', methods=['POST'])
@login_required
def publish_schedule():
    """
    Step 8: Publish the final schedule as the working schedule.
    Makes it visible to all users as the current schedule.
    """
    global current_schedule, published_schedule, planner_state

    if current_user.role not in ('admin', 'planner'):
        return jsonify({'error': 'Unauthorized'}), 403

    final = planner_state.get('final_schedule')
    if not final:
        return jsonify({'error': 'No final schedule to publish. Generate a final schedule first.'}), 400

    now = datetime.now()

    # Update the current_schedule (existing global state for backward compat)
    current_schedule = {
        'generated_at': now,
        'published_by': current_user.username,
        'active_mode': final['scenario_key'],
        'modes': {
            final['scenario_key']: {
                'orders': final['orders'],
                'serialized_orders': final['serialized_orders'],
                'stats': final['stats'],
                'reports': final['reports'],
            }
        },
        'orders': final['orders'],
        'baseline_orders': [],
        'reports': final['reports'],
        'stats': final['stats'],
        'serialized_orders': final['serialized_orders'],
    }

    # Update published schedule state
    published_schedule['schedule_data'] = final
    published_schedule['published_at'] = now
    published_schedule['published_by'] = current_user.username
    published_schedule['mode_label'] = final['scenario_label']

    # Persist to GCS — both the legacy state and the new published state
    gcs_storage.save_schedule_state({
        'generated_at': now.isoformat(),
        'published_by': current_user.username,
        'active_mode': final['scenario_key'],
        'modes': {
            final['scenario_key']: {
                'orders': final['serialized_orders'],
                'stats': final['stats'],
                'reports': final['reports'],
            }
        },
        'orders': final['serialized_orders'],
        'stats': final['stats'],
        'reports': final['reports'],
    })

    gcs_storage.save_published_schedule({
        'published_at': now.isoformat(),
        'published_by': current_user.username,
        'mode_label': final['scenario_label'],
        'scenario_key': final['scenario_key'],
        'stats': final['stats'],
        'orders': final['serialized_orders'],
        'reports': final['reports'],
        'approved_request_count': final.get('approved_request_count', 0),
    })

    # Clear planner workflow state
    planner_state['final_schedule'] = None
    planner_state['scenarios'] = {}
    planner_state['base_scenario'] = None
    planner_state['base_schedule'] = None

    # Mark processed special requests as complete
    all_requests = gcs_storage.load_special_requests()
    for req in all_requests:
        if req.get('status') == 'approved':
            req['status'] = 'published'
            req['published_at'] = now.isoformat()
    gcs_storage.save_special_requests(all_requests)

    # Auto-generate alerts on publish
    try:
        alerts_data = generate_alert_report(final['serialized_orders'])
        gcs_storage.save_alerts(alerts_data)
        print(f"[Publish] Generated alerts: {alerts_data['summary']}")
    except Exception as e:
        print(f"[Publish] Alert generation failed (non-blocking): {e}")

    # Create notification for schedule publish
    create_notification(
        'success',
        f'Schedule published by {current_user.username} ({final["scenario_label"]})',
        target_roles=None,
        related_entity={'type': 'schedule', 'value': now.isoformat()}
    )

    return jsonify({
        'success': True,
        'published_at': now.isoformat(),
        'published_by': current_user.username,
        'mode_label': final['scenario_label'],
        'stats': final['stats'],
    })


@app.route('/api/planner/status')
@login_required
def get_planner_status():
    """Get current planner workflow status."""
    if current_user.role not in ('admin', 'planner'):
        return jsonify({'error': 'Unauthorized'}), 403

    # Count pending requests
    all_requests = gcs_storage.load_special_requests()
    pending_count = sum(1 for r in all_requests if r.get('status') == 'pending')

    status = {
        'has_scenarios': bool(planner_state.get('scenarios')),
        'simulated_at': planner_state.get('simulated_at', '').isoformat() if planner_state.get('simulated_at') else None,
        'base_scenario': planner_state.get('base_scenario'),
        'base_scenario_label': SCENARIO_CONFIGS.get(planner_state.get('base_scenario', ''), {}).get('label'),
        'has_final_schedule': bool(planner_state.get('final_schedule')),
        'pending_request_count': pending_count,
        'published_at': published_schedule.get('published_at', '').isoformat() if published_schedule.get('published_at') else None,
        'published_by': published_schedule.get('published_by'),
        'published_mode': published_schedule.get('mode_label'),
    }

    return jsonify(status)


# ============== Error Handlers ==============

@app.errorhandler(401)
def unauthorized(e):
    """Handle unauthorized access."""
    if request.path.startswith('/api/'):
        return jsonify({'error': 'Unauthorized'}), 401
    return redirect(url_for('login'))


@app.errorhandler(404)
def not_found(e):
    """Handle 404 errors."""
    if request.path.startswith('/api/'):
        return jsonify({'error': 'Not found'}), 404
    return render_template('404.html'), 404


@app.errorhandler(500)
def server_error(e):
    """Handle 500 errors."""
    if request.path.startswith('/api/'):
        return jsonify({'error': 'Internal server error'}), 500
    return render_template('500.html'), 500


# ============== Notifications ==============


def create_notification(notif_type, message, target_roles=None, related_entity=None):
    """Create a notification and persist it.

    Args:
        notif_type: 'info', 'warning', 'success', 'danger'
        message: Short notification text
        target_roles: List of roles to show this to, or None for all
        related_entity: Optional dict like {'type': 'schedule', 'value': '...'}
    """
    notifications = gcs_storage.load_notifications()

    notif = {
        'id': f"NTF-{datetime.now().strftime('%Y%m%d%H%M%S')}-{len(notifications)}",
        'type': notif_type,
        'message': message,
        'target_roles': target_roles,
        'created_at': datetime.now().isoformat(),
        'read_by': [],
        'related_entity': related_entity,
    }

    notifications.append(notif)
    gcs_storage.save_notifications(notifications)
    return notif


@app.route('/api/notifications')
@login_required
def get_notifications():
    """Get notifications for the current user. Auto-marks old ones as read."""
    notifications = gcs_storage.load_notifications()
    now = datetime.now()
    changed = False

    user_notifs = []
    for n in notifications:
        # Filter by target role
        targets = n.get('target_roles')
        if targets and current_user.role not in targets:
            continue

        # Auto-mark as read after 7 days
        created = datetime.fromisoformat(n['created_at'])
        if (now - created).days >= 7 and current_user.username not in n.get('read_by', []):
            n.setdefault('read_by', []).append(current_user.username)
            changed = True

        is_read = current_user.username in n.get('read_by', [])
        user_notifs.append({
            'id': n['id'],
            'type': n['type'],
            'message': n['message'],
            'created_at': n['created_at'],
            'is_read': is_read,
            'related_entity': n.get('related_entity'),
        })

    if changed:
        gcs_storage.save_notifications(notifications)

    # Newest first, limit to 50
    user_notifs.reverse()
    unread_count = sum(1 for n in user_notifs if not n['is_read'])

    return jsonify({
        'notifications': user_notifs[:50],
        'unread_count': unread_count
    })


@app.route('/api/notifications/<notif_id>/read', methods=['POST'])
@login_required
def mark_notification_read(notif_id):
    """Mark a single notification as read."""
    notifications = gcs_storage.load_notifications()
    for n in notifications:
        if n['id'] == notif_id:
            if current_user.username not in n.get('read_by', []):
                n.setdefault('read_by', []).append(current_user.username)
                gcs_storage.save_notifications(notifications)
            return jsonify({'success': True})
    return jsonify({'error': 'Notification not found'}), 404


@app.route('/api/notifications/read-all', methods=['POST'])
@login_required
def mark_all_notifications_read():
    """Mark all notifications as read for current user."""
    notifications = gcs_storage.load_notifications()
    changed = False
    for n in notifications:
        targets = n.get('target_roles')
        if targets and current_user.role not in targets:
            continue
        if current_user.username not in n.get('read_by', []):
            n.setdefault('read_by', []).append(current_user.username)
            changed = True
    if changed:
        gcs_storage.save_notifications(notifications)
    return jsonify({'success': True})


@app.route('/notifications')
@login_required
def notifications_page():
    """Notifications page — full list with filtering."""
    return render_template('notifications.html')


# ============== Alert Reports ==============


def generate_alert_report(orders_data):
    """Generate alert report from serialized order data.

    Analyzes orders for: late orders, promise date risk, core shortage, machine utilization.

    Args:
        orders_data: List of serialized order dicts (from schedule state)

    Returns:
        Dict with alerts list and summary counts
    """
    now = datetime.now()
    alerts = []
    late_orders = []
    at_risk_orders = []
    core_usage = {}  # core_number -> count
    machine_usage = {}  # desma -> count

    AT_RISK_DAYS = 3  # Orders within 3 days of promise date
    MACHINE_OVERLOAD_PCT = 90
    MACHINE_UNDERLOAD_PCT = 40
    TOTAL_MACHINES = 5  # D1-D5

    for order in orders_data:
        wo = order.get('wo_number', '')

        # --- Late Order Summary ---
        if order.get('on_time_status') == 'Late':
            promise = order.get('promise_date', '')
            completion = order.get('completion_date', '')
            days_late = ''
            if promise and completion:
                try:
                    p = datetime.fromisoformat(promise)
                    c = datetime.fromisoformat(completion)
                    days_late = (c - p).days
                except (ValueError, TypeError):
                    pass
            late_orders.append({
                'wo_number': wo,
                'customer': order.get('customer', ''),
                'part_number': order.get('part_number', ''),
                'promise_date': promise,
                'completion_date': completion,
                'days_late': days_late,
            })

        # --- Promise Date Risk ---
        elif order.get('on_time_status') == 'At Risk':
            at_risk_orders.append({
                'wo_number': wo,
                'customer': order.get('customer', ''),
                'part_number': order.get('part_number', ''),
                'promise_date': order.get('promise_date', ''),
                'completion_date': order.get('completion_date', ''),
            })
        elif order.get('on_time_status') == 'On Time' and order.get('promise_date') and order.get('completion_date'):
            try:
                promise = datetime.fromisoformat(order['promise_date'])
                completion = datetime.fromisoformat(order['completion_date'])
                buffer_days = (promise - completion).days
                if 0 < buffer_days <= AT_RISK_DAYS:
                    at_risk_orders.append({
                        'wo_number': wo,
                        'customer': order.get('customer', ''),
                        'part_number': order.get('part_number', ''),
                        'promise_date': order['promise_date'],
                        'completion_date': order['completion_date'],
                        'buffer_days': buffer_days,
                    })
            except (ValueError, TypeError):
                pass

        # --- Core Shortage tracking ---
        core = order.get('core', '')
        if core:
            core_num = core.split('-')[0] if '-' in core else core
            core_usage[core_num] = core_usage.get(core_num, 0) + 1

        # --- Machine utilization tracking ---
        # We track from operations if available, otherwise from serialized data
        desma = order.get('planned_desma', '')
        if desma:
            machine_usage[desma] = machine_usage.get(desma, 0) + 1

    # Build alerts list
    if late_orders:
        alerts.append({
            'type': 'late_orders',
            'severity': 'danger',
            'title': 'Late Orders',
            'count': len(late_orders),
            'message': f'{len(late_orders)} order(s) past their promise date',
            'details': late_orders,
        })

    if at_risk_orders:
        alerts.append({
            'type': 'promise_date_risk',
            'severity': 'warning',
            'title': 'Promise Date Risk',
            'count': len(at_risk_orders),
            'message': f'{len(at_risk_orders)} order(s) at risk of missing promise date',
            'details': at_risk_orders,
        })

    # Core shortage: cores used by 3+ orders concurrently may indicate shortage
    overloaded_cores = {k: v for k, v in core_usage.items() if v >= 3}
    if overloaded_cores:
        alerts.append({
            'type': 'core_shortage',
            'severity': 'warning',
            'title': 'Core Utilization',
            'count': len(overloaded_cores),
            'message': f'{len(overloaded_cores)} core(s) assigned to 3+ orders',
            'details': [{'core': k, 'order_count': v} for k, v in sorted(overloaded_cores.items(), key=lambda x: -x[1])],
        })

    # Machine utilization
    total_orders = len(orders_data)
    if total_orders > 0 and machine_usage:
        avg_per_machine = total_orders / TOTAL_MACHINES
        overloaded = []
        underloaded = []
        for m in [f'D{i}' for i in range(1, TOTAL_MACHINES + 1)]:
            count = machine_usage.get(m, 0)
            pct = (count / total_orders * 100) if total_orders else 0
            if pct > MACHINE_OVERLOAD_PCT / TOTAL_MACHINES * 100:
                overloaded.append({'machine': m, 'order_count': count, 'pct': round(pct, 1)})
            elif count == 0 and total_orders > 0:
                underloaded.append({'machine': m, 'order_count': 0, 'pct': 0})

        machine_details = []
        for m in [f'D{i}' for i in range(1, TOTAL_MACHINES + 1)]:
            count = machine_usage.get(m, 0)
            pct = round(count / total_orders * 100, 1) if total_orders else 0
            machine_details.append({'machine': m, 'order_count': count, 'pct': pct})

        max_pct = max(d['pct'] for d in machine_details) if machine_details else 0
        min_pct = min(d['pct'] for d in machine_details) if machine_details else 0
        imbalance = max_pct - min_pct

        severity = 'info'
        if imbalance > 30:
            severity = 'warning'
        if imbalance > 50:
            severity = 'danger'

        alerts.append({
            'type': 'machine_utilization',
            'severity': severity,
            'title': 'Machine Utilization',
            'count': TOTAL_MACHINES,
            'message': f'Utilization spread: {min_pct}% - {max_pct}% across {TOTAL_MACHINES} machines',
            'details': machine_details,
        })

    summary = {
        'late_count': len(late_orders),
        'at_risk_count': len(at_risk_orders),
        'core_alerts': len(overloaded_cores) if 'overloaded_cores' in dir() else 0,
        'total_alerts': len(alerts),
    }

    return {
        'generated_at': now.isoformat(),
        'alerts': alerts,
        'summary': summary,
    }


@app.route('/api/alerts')
@login_required
def get_alerts():
    """Get current alert report."""
    alerts_data = gcs_storage.load_alerts()
    if not alerts_data:
        return jsonify({'alerts': [], 'summary': {'total_alerts': 0}, 'generated_at': None})
    return jsonify(alerts_data)


@app.route('/api/alerts/generate', methods=['POST'])
@login_required
def generate_alerts():
    """On-demand alert generation from current schedule data."""
    if current_user.role not in ('admin', 'planner'):
        return jsonify({'error': 'Unauthorized'}), 403

    # Get orders from published schedule or current schedule
    orders_data = []
    pub = gcs_storage.load_published_schedule()
    if pub and pub.get('orders'):
        orders_data = pub['orders']
    elif current_schedule.get('serialized_orders'):
        orders_data = current_schedule['serialized_orders']
    else:
        state = gcs_storage.load_schedule_state()
        if state and state.get('orders'):
            orders_data = state['orders']

    if not orders_data:
        return jsonify({'error': 'No schedule data available. Generate a schedule first.'}), 400

    alerts_data = generate_alert_report(orders_data)
    gcs_storage.save_alerts(alerts_data)

    # Create notification if there are new critical alerts
    late_count = alerts_data['summary'].get('late_count', 0)
    risk_count = alerts_data['summary'].get('at_risk_count', 0)
    if late_count > 0 or risk_count > 0:
        msg_parts = []
        if late_count:
            msg_parts.append(f'{late_count} late')
        if risk_count:
            msg_parts.append(f'{risk_count} at-risk')
        create_notification(
            'warning',
            f'Alert Report: {", ".join(msg_parts)} orders detected',
            target_roles=None,
            related_entity={'type': 'alert_report', 'value': alerts_data['generated_at']}
        )

    return jsonify(alerts_data)


@app.route('/alerts')
@login_required
def alerts_page():
    """Dedicated alerts page with full details and filtering."""
    can_generate = current_user.role in ('admin', 'planner')
    return render_template('alerts.html', can_generate=can_generate)


# ============== Main ==============

def run_development():
    """Run the development server."""
    host = os.environ.get('HOST', '0.0.0.0')
    port = int(os.environ.get('PORT', 5000))

    print("=" * 60)
    print("EstradaBot - Web Interface (Development)")
    print("=" * 60)
    print(f"Upload folder: {app.config['UPLOAD_FOLDER']}")
    print(f"Output folder: {app.config['OUTPUT_FOLDER']}")
    print(f"Starting server at http://{host}:{port}")
    print("=" * 60)
    print("WARNING: Using development server. For production, use:")
    print("  waitress-serve --port=5000 app:app")
    print("=" * 60)

    app.run(debug=True, host=host, port=port)


def run_production():
    """Run the production server with Waitress."""
    from waitress import serve

    host = os.environ.get('HOST', '0.0.0.0')
    port = int(os.environ.get('PORT', 5000))

    print("=" * 60)
    print("EstradaBot - Web Interface (Production)")
    print("=" * 60)
    print(f"Upload folder: {app.config['UPLOAD_FOLDER']}")
    print(f"Output folder: {app.config['OUTPUT_FOLDER']}")
    print(f"Starting Waitress server at http://{host}:{port}")
    print("=" * 60)

    serve(app, host=host, port=port, threads=4)


if __name__ == '__main__':
    env = os.environ.get('FLASK_ENV', 'development')

    if env == 'production':
        run_production()
    else:
        run_development()
