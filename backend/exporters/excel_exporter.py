"""
Excel Exporter
Export schedules and reports to Excel format.
"""

import pandas as pd
from datetime import datetime
from typing import List, Dict, Any
from pathlib import Path

# Add parent directory to path for imports
import sys
sys.path.insert(0, str(Path(__file__).parent.parent))


def _write_unscheduled_tab(writer, unscheduled_orders: List[Dict]) -> None:
    """Write the Unscheduled Orders tab to an open ExcelWriter."""
    if not unscheduled_orders:
        rows = [{'WO#': '(none)', 'Part Number': '', 'Description': '',
                 'Customer': '', 'Core#': '', 'Priority': '', 'Current Op#': '', 'Notes': ''}]
    else:
        rows = []
        for o in unscheduled_orders:
            rows.append({
                'WO#': o.get('wo_number', ''),
                'Part Number': o.get('part_number', ''),
                'Description': str(o.get('description', '') or '')[:60],
                'Customer': o.get('customer', '') or '',
                'Core#': o.get('core_number', '') or '',
                'Priority': o.get('priority', 'Normal'),
                'Current Op#': o.get('oso_op_number', '') or '',
                'Notes': o.get('unscheduled_reason', ''),
            })
    df = pd.DataFrame(rows)
    df.to_excel(writer, sheet_name='Unscheduled Orders', index=False)
    ws = writer.sheets['Unscheduled Orders']
    from openpyxl.utils import get_column_letter
    for idx, col in enumerate(df.columns):
        col_data = df[col].fillna('').astype(str)
        max_len = max(col_data.str.len().max() if len(col_data) > 0 else 0, len(col)) + 2
        ws.column_dimensions[get_column_letter(idx + 1)].width = min(max_len, 50)
    ws.freeze_panes = 'A2'


def export_master_schedule(scheduled_orders: List, output_path: str,
                           unscheduled_orders: List[Dict] = None) -> str:
    """
    Export the master schedule to Excel.

    Args:
        scheduled_orders: List of ScheduledOrder objects
        output_path: Path for output Excel file

    Returns:
        Path to the created file
    """
    # Convert to list of dictionaries
    data = []
    for order in scheduled_orders:
        data.append({
            'WO#': order.wo_number,
            'Serial Number': getattr(order, 'serial_number', None) or '',
            'Part Number': order.part_number,
            'Description': order.description[:50] if order.description else '',
            'Customer': order.customer,
            'Type': 'Reline' if order.is_reline else 'New',
            'Core': order.assigned_core,
            'Rubber Type': order.rubber_type,
            'WO Creation Date': order.creation_date,
            'Blast Date': order.blast_date,
            'Completion Date': order.completion_date,
            'Turnaround (days)': order.turnaround_days,
            'Basic Finish Date': getattr(order, 'basic_finish_date', None),
            'Promise Date': order.promise_date,
            'On-Time': 'Yes' if order.on_time else 'No',
            'Days Idle': getattr(order, 'days_idle', None) if getattr(order, 'days_idle', None) is not None else '',
            'Special Instructions': getattr(order, 'special_instructions', '') or ''
        })

    df = pd.DataFrame(data)

    # Create Excel writer with formatting
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Master Schedule', index=False)

        # Get workbook and worksheet
        workbook = writer.book
        worksheet = writer.sheets['Master Schedule']

        # Auto-adjust column widths
        for idx, col in enumerate(df.columns):
            # Convert to string and get max length, handling NaN
            col_data = df[col].fillna('').astype(str)
            max_data_len = col_data.str.len().max() if len(col_data) > 0 else 0
            max_length = max(max_data_len, len(col)) + 2
            # Handle columns beyond Z (use openpyxl's get_column_letter)
            from openpyxl.utils import get_column_letter
            worksheet.column_dimensions[get_column_letter(idx + 1)].width = min(max_length, 40)

        # Freeze top row
        worksheet.freeze_panes = 'A2'

        # Unscheduled Orders tab
        if unscheduled_orders is not None:
            _write_unscheduled_tab(writer, unscheduled_orders)

    print(f"[OK] Master schedule exported to: {output_path}")
    return output_path


def export_blast_schedule(scheduled_orders: List, output_path: str,
                          reorder_sequence: List = None,
                          currently_blasting: List = None,
                          unscheduled_orders: List[Dict] = None) -> str:
    """
    Export the BLAST operation schedule (printable).
    If reorder_sequence is provided, orders are sorted by that sequence
    instead of by blast_date (manual override by planner).
    If currently_blasting is provided, those WIP orders (already on the blaster)
    are prepended at the top with status 'IN PROGRESS'.
    """
    orders_with_blast = [o for o in scheduled_orders if o.blast_date]

    if reorder_sequence:
        # Apply custom reorder: build index by WO#, then sort by sequence position
        wo_index = {o.wo_number: o for o in orders_with_blast}
        reordered = []
        for wo in reorder_sequence:
            if wo in wo_index:
                reordered.append(wo_index.pop(wo))
        # Append any remaining orders not in the reorder sequence
        for o in orders_with_blast:
            if o.wo_number in wo_index:
                reordered.append(o)
        orders_with_blast = reordered
    else:
        # Default: sort by BLAST date
        orders_with_blast.sort(key=lambda x: x.blast_date)

    data = []

    # Prepend currently-blasting WIP orders at the top
    if currently_blasting:
        for seq, wip in enumerate(currently_blasting, 1):
            op_start = wip.get('operation_start_date')
            blast_date_str = op_start.strftime('%m/%d/%Y') if op_start else 'IN PROGRESS'
            blast_time_str = op_start.strftime('%H:%M') if op_start else ''
            row = {
                'Seq': f'WIP-{seq}',
                'WO#': wip.get('wo_number', ''),
                'Part Number': wip.get('part_number', ''),
                'Description': str(wip.get('description', ''))[:50] if wip.get('description') else '',
                'Customer': '',
                'Blast Date': blast_date_str,
                'Blast Time': blast_time_str,
                'Core Required': '',
                'Supermarket Location': '',
                'Special Instructions': 'Currently on blaster',
                'Planned Desma': ''
            }
            data.append(row)

    for seq, order in enumerate(orders_with_blast, 1):
        row = {
            'Seq': seq,
            'WO#': order.wo_number,
            'Part Number': order.part_number,
            'Description': str(order.description)[:50] if order.description else '',
            'Customer': order.customer[:30] if order.customer else '',
            'Blast Date': order.blast_date.strftime('%m/%d/%Y') if order.blast_date else '',
            'Blast Time': order.blast_date.strftime('%H:%M') if order.blast_date else '',
            'Core Required': order.assigned_core,
            'Supermarket Location': getattr(order, 'supermarket_location', '') or '',
            'Special Instructions': getattr(order, 'special_instructions', '') or '',
            'Planned Desma': getattr(order, 'planned_desma', '') or '',
            'Op#': getattr(order, 'oso_op_number', '') or '',
            'Current Op Description': getattr(order, 'oso_op_description', '') or ''
        }
        if reorder_sequence:
            row['Manual Override'] = 'Yes'
        data.append(row)

    df = pd.DataFrame(data)

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='BLAST Schedule', index=False)

        worksheet = writer.sheets['BLAST Schedule']

        # Format for printing
        from openpyxl.utils import get_column_letter
        for idx, col in enumerate(df.columns):
            col_data = df[col].fillna('').astype(str)
            max_data_len = col_data.str.len().max() if len(col_data) > 0 else 0
            max_length = max(max_data_len, len(col)) + 2
            worksheet.column_dimensions[get_column_letter(idx + 1)].width = min(max_length, 30)

        worksheet.freeze_panes = 'A2'

        # Unscheduled Orders tab
        if unscheduled_orders is not None:
            _write_unscheduled_tab(writer, unscheduled_orders)

    print(f"[OK] BLAST schedule exported to: {output_path}")
    return output_path


def export_core_schedule(scheduled_orders: List, output_path: str) -> str:
    """
    Export the Core Oven loading schedule.
    """
    # Extract core loading times (based on BLAST - 2.5 hours for heating)
    core_loads = []

    for order in scheduled_orders:
        if order.blast_date and order.assigned_core:
            # Core needs to be loaded ~2.5 hours before assembly
            # Assembly is after BLAST, so approximate
            from datetime import timedelta
            core_load_time = order.blast_date - timedelta(hours=1)  # Approximation

            core_loads.append({
                'Core': order.assigned_core,
                'Special Instructions': getattr(order, 'special_instructions', '') or '',
                'Load Date': core_load_time.strftime('%Y-%m-%d'),
                'Load Time': core_load_time.strftime('%H:%M'),
                'For WO#': order.wo_number,
                'Part Number': order.part_number,
                'Description': str(order.description)[:50] if order.description else ''
            })

    # Sort by load time
    core_loads.sort(key=lambda x: (x['Load Date'], x['Load Time']))

    # Add sequence numbers
    for seq, load in enumerate(core_loads, 1):
        load['Seq'] = seq

    # Reorder columns
    df = pd.DataFrame(core_loads)
    if not df.empty:
        df = df[['Seq', 'Core', 'Special Instructions', 'Load Date', 'Load Time', 'For WO#', 'Part Number', 'Description']]

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Core Oven Schedule', index=False)

        worksheet = writer.sheets['Core Oven Schedule']

        from openpyxl.utils import get_column_letter
        for idx, col in enumerate(df.columns):
            col_data = df[col].fillna('').astype(str)
            max_data_len = col_data.str.len().max() if len(col_data) > 0 else 0
            max_length = max(max_data_len, len(col)) + 2
            worksheet.column_dimensions[get_column_letter(idx + 1)].width = min(max_length, 25)

        worksheet.freeze_panes = 'A2'

    print(f"[OK] Core oven schedule exported to: {output_path}")
    return output_path


def export_pending_core_report(pending_orders: List[Dict], output_path: str) -> str:
    """
    Export the Pending Core report - orders that cannot be scheduled
    because their required core is not in inventory.

    Args:
        pending_orders: List of order dicts with pending core info
        output_path: Path for output Excel file

    Returns:
        Path to the created file
    """
    data = []
    for order in pending_orders:
        # Format dates safely
        created_on = order.get('created_on')
        if created_on and hasattr(created_on, 'strftime'):
            created_on_str = created_on.strftime('%Y-%m-%d')
        else:
            created_on_str = str(created_on) if created_on else ''

        promise_date = order.get('promise_date')
        if promise_date and hasattr(promise_date, 'strftime'):
            promise_date_str = promise_date.strftime('%Y-%m-%d')
        else:
            promise_date_str = str(promise_date) if promise_date else ''

        data.append({
            'WO#': order.get('wo_number', ''),
            'Part Number': order.get('part_number', ''),
            'Description': (order.get('description', '') or '')[:50],
            'Customer': order.get('customer', ''),
            'Core Number Needed': order.get('core_number_needed', 'Unknown'),
            'Reason': order.get('reason', 'Core not in inventory'),
            'Created On': created_on_str,
            'Promise Date': promise_date_str
        })

    df = pd.DataFrame(data)

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Pending Core', index=False)

        worksheet = writer.sheets['Pending Core']

        # Format columns
        from openpyxl.utils import get_column_letter
        for idx, col in enumerate(df.columns):
            col_data = df[col].fillna('').astype(str)
            max_data_len = col_data.str.len().max() if len(col_data) > 0 else 0
            max_length = max(max_data_len, len(col)) + 2
            worksheet.column_dimensions[get_column_letter(idx + 1)].width = min(max_length, 40)

        worksheet.freeze_panes = 'A2'

    print(f"[OK] Pending core report exported to: {output_path}")
    return output_path


def export_all_reports(scheduler, output_dir: str = None) -> Dict[str, str]:
    """
    Export all reports from a scheduler instance.

    Args:
        scheduler: ProductionScheduler instance with scheduled_orders
        output_dir: Output directory path. Defaults to project's outputs folder.

    Returns:
        Dictionary of report names to file paths
    """
    from pathlib import Path

    if output_dir is None:
        # Default to project root's outputs folder
        project_root = Path(__file__).parent.parent.parent
        output_dir = project_root / "outputs"
    else:
        output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')

    files = {}

    # Master Schedule
    master_path = output_dir / f"Master_Schedule_{timestamp}.xlsx"
    files['master_schedule'] = export_master_schedule(
        scheduler.scheduled_orders, str(master_path)
    )

    # BLAST Schedule
    blast_path = output_dir / f"BLAST_Schedule_{timestamp}.xlsx"
    files['blast_schedule'] = export_blast_schedule(
        scheduler.scheduled_orders, str(blast_path)
    )

    # Core Oven Schedule
    core_path = output_dir / f"Core_Oven_Schedule_{timestamp}.xlsx"
    files['core_schedule'] = export_core_schedule(
        scheduler.scheduled_orders, str(core_path)
    )

    # Pending Core Report (only if there are pending orders)
    if hasattr(scheduler, 'pending_core_orders') and scheduler.pending_core_orders:
        pending_path = output_dir / f"Pending_Core_{timestamp}.xlsx"
        files['pending_core'] = export_pending_core_report(
            scheduler.pending_core_orders, str(pending_path)
        )

    print(f"\n[OK] All reports exported to: {output_dir}")

    return files


if __name__ == "__main__":
    # Test export with DES scheduler
    import sys
    from pathlib import Path

    # Add parent to path
    sys.path.insert(0, str(Path(__file__).parent.parent))

    from data_loader import DataLoader
    from algorithms.des_scheduler import DESScheduler
    from exporters.impact_analysis_exporter import generate_impact_analysis

    print("Testing Excel Export with DES Scheduler")
    print("=" * 60)

    # Load data
    loader = DataLoader()
    if not loader.load_all():
        print("Failed to load data")
        sys.exit(1)

    # Load hot list (optional)
    loader.load_hot_list()

    from datetime import datetime
    start_date = datetime(2026, 2, 2, 5, 30)  # Monday Feb 2, 2026 at 5:30 AM (after handover)

    # Create and run baseline scheduler (without hot list) for impact analysis
    baseline_scheduler = None
    if loader.hot_list_entries:
        print("\n" + "=" * 60)
        print("BASELINE SCHEDULE (without hot list)")
        print("=" * 60)

        baseline_scheduler = DESScheduler(
            orders=loader.orders,
            core_mapping=loader.core_mapping,
            core_inventory=loader.core_inventory,
            operations=loader.operations
        )
        baseline_scheduler.schedule_orders(start_date=start_date)

    # Create and run DES scheduler (with hot list)
    print("\n" + "=" * 60)
    print("MAIN SCHEDULE (with hot list prioritization)")
    print("=" * 60)

    scheduler = DESScheduler(
        orders=loader.orders,
        core_mapping=loader.core_mapping,
        core_inventory=loader.core_inventory,
        operations=loader.operations
    )

    scheduler.schedule_orders(
        start_date=start_date,
        hot_list_entries=loader.hot_list_entries
    )

    # Print summary
    scheduler.print_summary()

    # Export all reports
    files = export_all_reports(scheduler)

    # Generate impact analysis if we have hot list and baseline
    if loader.hot_list_entries and baseline_scheduler:
        project_root = Path(__file__).parent.parent.parent
        output_dir = project_root / "outputs"
        impact_file = generate_impact_analysis(
            scheduled_orders=scheduler.scheduled_orders,
            baseline_orders=baseline_scheduler.scheduled_orders,
            hot_list_entries=loader.hot_list_entries,
            hot_list_core_shortages=getattr(scheduler, 'hot_list_core_shortages', []),
            output_dir=str(output_dir)
        )
        files['impact_analysis'] = impact_file

    print("\n[OK] Export complete!")
    for name, path in files.items():
        print(f"   {name}: {path}")
